#!/usr/bin/env python3
"""C2 probe — three ideas for getting geodata on chip suppliers.

A class C2 source names a mill, not a harvest area. Harmac's record stops at
the mill gate, so there is nothing to look up. This script tests three ways
round that, and reports honestly on which of them actually work.

    1  OPERATOR ENVELOPE
       Most BC chip suppliers are licensees in their own right. Find their
       FTEN client number, pull their cut blocks, narrow by district and
       delivery window. Not an attribution - it says the fibre came from
       somewhere in these blocks - but it is mapped, public and defensible,
       and it turns "send us everything" into "which of these forty".

    2  HBS BY CLIENT   -- SETTLED, NO
       Tested 12 Aug 2026. The public screen takes clientNumber as the timber
       mark, truncates it to six characters and rejects it: "Timber mark
       invalid, must be in Timber Mark table." There is no public client-keyed
       search. Left in the script so the answer stays checkable, and so nobody
       spends an afternoon rediscovering it.

    3  THREE LAYERS
       793  Major Timber Processing Facilities - 252 mills as points, with
            annual capacity in the same units as Harmac's delivery data.
       814  Fibre Recovery Zones - 664 polygons, and small: the samples are
            1.8 and 8.3 hectares. Probably roadside residual pockets rather
            than supply catchments. --test-814 checks whether any resolved
            harvest area actually falls inside one.
       543  Harvested Areas (consolidated cutblocks) - 592,776 polygons with
            a harvest year, built from forest cover, RESULTS, tenure
            applications AND satellite change detection.

       543 is the interesting one. It is documented as crown lands only, but
       change detection does not respect tenure boundaries, so the
       documentation and the data may not agree. If it does cover private
       land it is the first public source of actual harvest polygons for
       class B - not an ownership catchment but an observed harvest.
       --test-543 answers that empirically.

Usage
-----
    python tools/c2_probe.py --suppliers "Coastland Wood Industries,Interfor"
    python tools/c2_probe.py --register HPA1_..._v5.xlsx --class C2
    python tools/c2_probe.py --layers-only
    python tools/c2_probe.py --suppliers "Interfor" --district DSI --since 2025-07-01

Nothing here is production code. It is a way of finding out whether an idea is
worth building, and it prints what it did so the answer can be checked.

Requires: requests   (openpyxl only for --register)
Licence : Contains information licensed under the Open Government Licence -
          British Columbia.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from datetime import datetime

import requests

ROOT = "https://delivery.maps.gov.bc.ca/arcgis/rest/services/mpcm/bcgwpub/MapServer"
BLOCKS = ROOT + "/340/query"
HBS_ROOT = "https://a100.gov.bc.ca/pub/hbs"

PAGE = 1000
TIMEOUT = 120
ATTRIBUTION = ("Contains information licensed under the "
               "Open Government Licence - British Columbia.")

S = requests.Session()
S.headers.update({"User-Agent": "NGIS-HARP-c2-probe/1.0"})


def post(url: str, params: dict) -> dict:
    r = S.post(url, data=params, timeout=TIMEOUT)
    r.raise_for_status()
    data = r.json()
    if "error" in data:
        raise RuntimeError(data["error"].get("message", str(data["error"])))
    return data


def q(v) -> str:
    return str(v).replace("'", "''")


# ───────────────────────── 1. operator envelope ────────────────────────────
#
# Name matching is tiered. A single generic word matches hundreds of unrelated
# holders - CEDAR returned the same 37 clients for three different suppliers -
# so the full name is tried first and a generic word only as a last resort,
# labelled as such.

CORP = {"LTD", "LTD.", "INC", "INC.", "LIMITED", "CORP", "CORP.",
        "CORPORATION", "COMPANY", "CO", "CO.", "LP", "LLP", "GP", "THE"}
GENERIC = {"FOREST", "FORESTRY", "LOGGING", "LUMBER", "TIMBER", "WOOD",
           "WOODS", "FIBRE", "FIBER", "MILL", "MILLS", "SAWMILL", "SAWMILLS",
           "PRODUCTS", "INDUSTRIES", "INDUSTRIAL", "ENTERPRISES", "CEDAR",
           "VALLEY", "ISLAND", "RIVER", "COAST", "COASTAL",
           "WEST", "EAST", "NORTH", "SOUTH", "SHAKE", "SHINGLE", "GROUP"}


# Words carrying no company identity: what an industry shares, corporate
# suffixes, and geography that qualifies a name without distinguishing it.
# A word here can never carry a match on its own.
NOISE = {
    # PACIFIC is not noise. It distinguishes NICOLA PACIFIC FOREST PRODUCTS
    # from NICOLA POST & RAIL, and treating it as filler matched a supplier to
    # a company five times its size.
    "CANADA", "CANADIAN", "BC", "B.C.", "BRITISH", "COLUMBIA",
    "WEST", "WESTERN", "EAST", "EASTERN", "NORTH", "NORTHERN", "SOUTH",
    "SOUTHERN", "COAST", "COASTAL", "ISLAND", "VALLEY", "GROUP", "HOLDINGS",
    "PLANING", "PLANER", "MANUFACTURING", "MFG", "OPERATIONS", "DIVISION",
    "PULP", "PAPER", "SHINGLE", "SHINGLES", "PLYWOOD", "VENEER", "CHIP",
    "CHIPS", "LOG", "LOGS", "SAW", "POST", "RAIL", "SPLIT", "BARGE",
}


# Abbreviations a company writes on one record and spells out on another.
# 'Coastland Wood Ind.' and 'COASTLAND WOOD INDUSTRIES LTD' are one firm; a
# matcher that cannot expand IND refuses a correct answer.
ABBREV = {
    "IND": "INDUSTRIES", "INDS": "INDUSTRIES", "BROS": "BROTHERS",
    "BRO": "BROTHERS", "MFG": "MANUFACTURING", "ENT": "ENTERPRISES",
    "ENTS": "ENTERPRISES", "PROD": "PRODUCTS", "PRODS": "PRODUCTS",
    "MTN": "MOUNTAIN", "CONST": "CONSTRUCTION", "CONTR": "CONTRACTING",
    "RES": "RESOURCES", "SVCS": "SERVICES", "SVC": "SERVICES",
}


def canon(name: str) -> str:
    """A company name reduced to something comparable.

    '&' and 'AND' are the same word here - 'Nicola Post and Rail' and
    'NICOLA POST & RAIL LTD' are one company, and a matcher that cannot see
    that is refusing a correct answer.
    """
    n = re.sub(r"\(.*?\)", " ", str(name or "")).upper()
    n = n.replace("&", " AND ")
    n = re.sub(r"[^A-Z0-9\s'-]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return " ".join(ABBREV.get(w, w) for w in n.split())


def identifying_words(name: str) -> list[str]:
    """The words that actually identify a company.

    Everything an industry shares is stripped: LUMBER, TIMBER, CEDAR, FOREST,
    SAWMILL and the rest, along with the corporate suffixes. What remains is
    what makes the name that firm rather than another one.
    """
    words = [w for w in canon(name).split() if w not in CORP]
    return [w for w in words
            if w not in GENERIC and w not in NOISE and w != "AND"
            and len(w) >= 3]


def verify(supplier: str, client: str) -> str:
    """Is this FTEN client actually the supplier? Returns a tier, or "".

    Searching finds candidates; this decides. The previous version had no such
    step - it searched for one distinctive word and accepted whatever came
    back, which is how "Imperial Fibre" reached Imperial Oil Resources and
    "Ludwig Lumber" reached Dunkley, Galloway and Kalesnikoff.

    A candidate now has to carry EVERY identifying word in the supplier's
    name. Ludwig matches Ludwig; it does not match Dunkley because they happen
    to share the word LUMBER.
    """
    sup, cli = canon(supplier), canon(client)
    if not sup or not cli:
        return ""
    if sup == cli:
        return "exact"

    words = identifying_words(supplier)
    if not words:
        # nothing but generic words - a name like "Valley Cedar" cannot be
        # verified against anything, so it is refused rather than guessed
        return ""

    cw = set(cli.split())

    def present(w, pool):
        # whole word, or a near-identical form - plurals and possessives, not
        # prefixes. Allowing a loose prefix made ALTAGAS match ALTA.
        if w in pool:
            return True
        return any(abs(len(c) - len(w)) <= 2 and
                   (c.startswith(w) or w.startswith(c)) for c in pool)

    for w in words:
        if not present(w, cw):
            return ""

    # And the other way. Every word that identifies the CLIENT must be
    # accounted for in the supplier's name too, or they are different
    # companies that happen to share a word - IMPERIAL FIBRE against IMPERIAL
    # OIL RESOURCES passed a one-directional test and is plainly wrong.
    client_words = identifying_words(client)
    sw = set(sup.split())
    unexplained = [w for w in client_words if not present(w, sw)]
    if unexplained:
        # The supplier's words are all here, but the client carries an
        # identifying word of its own - 'Gorman Group' against 'GORMAN BROS.
        # LUMBER'. They may well be the same family firm, and they may not.
        # Reported for someone to confirm, never used as a match.
        return "possible"

    # A name resting on one identifying word is inherently weak - STAR, DELTA,
    # IMPERIAL and CAPE all belong to several unrelated firms. Require the
    # client name to actually begin with that word, or it is only plausible.
    # 'Star Lumber' should not match NORTH STAR PLANING: the company is North
    # Star, not Star.
    if len(words) == 1:
        w = words[0]
        first = cli.split()[0] if cli.split() else ""
        if not (first == w or (abs(len(first) - len(w)) <= 2 and
                               (first.startswith(w) or w.startswith(first)))):
            return "possible"

    extra = len(cw) - len(sw)
    return "high" if extra <= 3 else "medium"


def name_tiers(name: str) -> list[tuple[str, list[str]]]:
    n = re.sub(r"\(.*?\)", " ", str(name or "")).upper()
    n = re.sub(r"[^A-Z0-9&\s'-]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    if not n:
        return []
    words = [w for w in n.split() if w not in CORP]
    distinct = [w for w in words if w not in GENERIC and len(w) >= 4]
    tiers = []
    if len(n) >= 5:
        tiers.append(("high (full name)", [n]))
    if len(words) >= 2:
        tiers.append(("high (leading words)", [" ".join(words[:2])]))
    if distinct:
        tiers.append(("medium (distinctive word)", distinct[:3]))
    rest = [w for w in words if w not in distinct and len(w) >= 3]
    if rest:
        tiers.append(("low (generic word)", rest[:3]))
    return tiers


# Matches that cannot be confirmed from the name alone. Reported at the end
# for a person to accept or reject; never used.
POSSIBLE: dict[str, list[str]] = {}


def find_clients(name: str) -> tuple[str, list[dict]]:
    """FTEN tenure holders that are actually this supplier.

    Search wide, verify hard. Every identifying word in the supplier's name
    must appear in the client's, so a shared industry word cannot carry a
    match on its own. Anything that fails verification is dropped rather than
    returned at a lower confidence - a wrong holder is thousands of blocks of
    someone else's tenure, and a declaration built on it would be wrong rather
    than merely broad.
    """
    words = identifying_words(name)
    if not words:
        return "", []

    # candidates: search on the full name and on each identifying word
    terms = [canon(name)] + words[:4]
    found: dict[tuple, dict] = {}
    for t in terms:
        if len(t) < 3:
            continue
        try:
            data = post(BLOCKS, {
                "where": "CLIENT_NAME LIKE '%{}%'".format(q(t).upper()),
                "outFields": "CLIENT_NAME,CLIENT_NUMBER,CLIENT_LOCATION_CODE",
                "returnDistinctValues": "true", "returnGeometry": "false",
                "resultRecordCount": 200, "f": "json"})
        except Exception:
            continue
        for f in data.get("features", []):
            a = f["attributes"]
            if not a.get("CLIENT_NAME"):
                continue
            key = (a["CLIENT_NAME"], a.get("CLIENT_NUMBER"),
                   a.get("CLIENT_LOCATION_CODE"))
            found[key] = {
                "client_name": a["CLIENT_NAME"].strip(),
                "client_number": (a.get("CLIENT_NUMBER") or "").strip(),
                "location": (a.get("CLIENT_LOCATION_CODE") or "").strip()}

    kept, possible, best = [], [], ""
    for c in found.values():
        tier = verify(name, c["client_name"])
        if not tier:
            continue
        if tier == "possible":
            # keep the number - a plausible match that cannot be keyed cannot
            # be ruled on, and a proposal nobody can accept is just noise
            possible.append({"client_name": c["client_name"],
                             "client_number": c["client_number"]})
            continue
        c["verified"] = tier
        kept.append(c)
        if tier == "exact" or (tier == "high" and best != "exact"):
            best = tier
        elif not best:
            best = tier

    if possible:
        POSSIBLE.setdefault(name, []).extend(possible)
    if not kept:
        return "", []
    label = {"exact": "high (exact name)", "high": "high (all words)",
             "medium": "medium (all words, broader client)"}.get(best, best)
    kept.sort(key=lambda c: {"exact": 0, "high": 1, "medium": 2}[c["verified"]])
    return label, kept


def count(where: str) -> int:
    try:
        return post(BLOCKS, {"where": where, "returnCountOnly": "true",
                             "f": "json"}).get("count", 0)
    except Exception:
        return 0


def envelope(client_number: str, location: str = "", district: str = "",
             since: str = "") -> dict:
    """How many blocks, and how much area, this holder has - and how much the
    district and date filters actually narrow it.

    The narrowing is the whole point. An envelope covering a supplier's entire
    provincial tenure is not useful; the same envelope cut to one district and
    one delivery year might be.
    """
    base = "CLIENT_NUMBER = '{}'".format(q(client_number))
    if location:
        base += " AND CLIENT_LOCATION_CODE = '{}'".format(q(location))

    out = {"client_number": client_number, "location": location,
           "blocks_all": count(base)}

    w = base
    if district:
        w += " AND GEOGRAPHIC_DISTRICT_CODE = '{}'".format(q(district))
        out["blocks_district"] = count(w)
    if since:
        w2 = w + (" AND DISTURBANCE_START_DATE > DATE '{}'".format(q(since)))
        out["blocks_window"] = count(w2)
        w = w2

    out["where"] = w
    try:
        data = post(BLOCKS, {"where": w, "outFields": "FEATURE_AREA",
                             "returnGeometry": "false",
                             "resultRecordCount": PAGE, "f": "json"})
        rows = [f["attributes"] for f in data.get("features", [])]
        out["area_ha"] = round(
            sum(r.get("FEATURE_AREA") or 0 for r in rows) / 10000.0, 1)
        out["area_truncated"] = len(rows) >= PAGE
    except Exception as exc:
        out["area_ha"] = 0
        out["error"] = str(exc)
    return out


# ────────────────────────── 2. HBS by client ───────────────────────────────

HBS_CANDIDATES = [
    ("timber mark query (known good)", "/opq/timberMarkQuery.do",
     {"pageName": "P480", "timberMark": "EDRWD"}),
    ("search for timber mark form", "/opq/P480.jsp", {}),
    ("client query", "/opq/clientQuery.do", {"clientNumber": "00158809"}),
    ("client search form", "/opq/P470.jsp", {}),
    ("marks by client", "/opq/timberMarkQuery.do",
     {"pageName": "P480", "clientNumber": "00158809"}),
    ("public home", "/home.jsp", {}),
]


def probe_hbs() -> list[dict]:
    """Does the public HBS expose anything keyed on client rather than mark?

    Guessing at URLs is crude, but the alternative is asserting that it does
    not exist without having looked. Every attempt is reported with what came
    back, so a hit can be followed up and a miss is not mistaken for proof.
    """
    out = []
    for label, path, params in HBS_CANDIDATES:
        url = HBS_ROOT + path
        row = {"probe": label, "url": url, "params": json.dumps(params)}
        try:
            r = S.get(url, params=params, timeout=60)
            body = r.text or ""
            row["status"] = r.status_code
            row["bytes"] = len(body)
            low = body.lower()
            row["looks_like_login"] = any(
                w in low for w in ("bceid", "sign in", "log in", "userid",
                                   "password"))
            row["has_client_field"] = "client" in low
            row["has_timber_mark"] = "timber mark" in low
            row["title"] = (re.search(r"<title>(.*?)</title>", body,
                                      re.I | re.S).group(1).strip()[:70]
                            if "<title>" in low else "")
        except Exception as exc:
            row["status"] = "error"
            row["error"] = str(exc)[:90]
        out.append(row)
        time.sleep(0.5)
    return out


# ──────────────────────────── 3. two layers ────────────────────────────────

LAYERS = {793: "Major Timber Processing Facilities",
          814: "Fibre Recovery Zones",
          543: "Harvested Areas (consolidated cutblocks)",
          238: "Generalized Forest Cover Ownership",
          748: "Natural Resource Districts"}

# Ownership classes that mean the land is not Crown provincial. Matched against
# whatever descriptive field layer 238 turns out to carry - the field names are
# read at run time rather than assumed.
PRIVATE_WORDS = ("PRIVATE", "CROWN GRANT", "FEE SIMPLE", "MUNICIPAL",
                 "INDIAN RESERVE", "FEDERAL")


def layer_url(layer: int) -> str:
    return "{}/{}".format(ROOT, layer)


def schema(layer: int) -> list[dict]:
    try:
        return post(layer_url(layer), {"f": "json"}).get("fields", [])
    except Exception:
        return []


def field_like(layer: int, *words: str) -> str:
    """First string field whose name contains all of these words."""
    for f in schema(layer):
        if f.get("type") != "esriFieldTypeString":
            continue
        u = f["name"].upper()
        if all(w in u for w in words):
            return f["name"]
    return ""


def distinct(layer: int, field: str, limit: int = 200) -> list[str]:
    try:
        data = post(layer_url(layer) + "/query",
                    {"where": "1=1", "outFields": field,
                     "returnDistinctValues": "true", "returnGeometry": "false",
                     "resultRecordCount": limit, "f": "json"})
    except Exception:
        return []
    out = []
    for f in data.get("features", []):
        v = (f["attributes"].get(field) or "").strip()
        if v and v not in out:
            out.append(v)
    return sorted(out)


def geometry_of(layer: int, where: str, limit: int = 1) -> list[dict]:
    try:
        data = post(layer_url(layer) + "/query",
                    {"where": where, "outFields": "*", "returnGeometry": "true",
                     "outSR": 4326, "resultRecordCount": limit, "f": "geojson"})
    except Exception:
        return []
    return data.get("features", []) or []


def count_within(layer: int, geom: dict, where: str = "1=1") -> int:
    """How many features of `layer` intersect this geometry. Server-side."""
    try:
        return post(layer_url(layer) + "/query", {
            "where": where, "geometry": json.dumps(geom),
            "geometryType": "esriGeometryPolygon",
            "spatialRel": "esriSpatialRelIntersects",
            "inSR": 4326, "returnCountOnly": "true", "f": "json"}).get("count", 0)
    except Exception:
        return -1


def sample_within(layer: int, geom: dict, fields: str, where: str = "1=1",
                  n: int = 5) -> list[dict]:
    try:
        data = post(layer_url(layer) + "/query", {
            "where": where, "geometry": json.dumps(geom),
            "geometryType": "esriGeometryPolygon",
            "spatialRel": "esriSpatialRelIntersects",
            "inSR": 4326, "outFields": fields, "returnGeometry": "false",
            "resultRecordCount": n, "f": "json"})
    except Exception:
        return []
    return [f["attributes"] for f in data.get("features", [])]


def bbox_polygon(features: list[dict]) -> dict | None:
    """A bounding box around a set of GeoJSON features."""
    xs, ys = [], []

    def walk(c):
        if isinstance(c, (int, float)):
            return
        if c and isinstance(c[0], (int, float)):
            xs.append(c[0]); ys.append(c[1])
            return
        for x in c:
            walk(x)

    for f in features:
        g = f.get("geometry") or {}
        if g.get("coordinates"):
            walk(g["coordinates"])
    if not xs:
        return None
    return {"rings": [[[min(xs), min(ys)], [max(xs), min(ys)],
                       [max(xs), max(ys)], [min(xs), max(ys)],
                       [min(xs), min(ys)]]]}


def test_543_private(district_code: str = "DSI",
                     since_year: int = 2020, parcels: int = 8) -> dict:
    """Does layer 543 hold harvests on private land?

    Documented as crown lands only, but it is built partly from satellite
    change detection, which has no way of knowing where a tenure boundary is.
    So rather than trust the description: find land that layer 238 classifies
    as private, and ask 543 whether it has any harvests there.

    If it does, class B gains its first public source of observed harvest
    polygons - not an ownership catchment, an actual cut.
    """
    out: dict = {"district": district_code, "since_year": since_year}

    dfield = ""
    for cand in ("DISTRICT_CODE", "ORG_UNIT", "ORG_UNIT_CODE"):
        if any(f["name"] == cand for f in schema(748)):
            dfield = cand
            break
    if not dfield:
        out["error"] = "no district code field on layer 748"
        return out
    dist = geometry_of(748, "{} = '{}'".format(dfield, q(district_code)))
    if not dist:
        out["error"] = "district {} not found".format(district_code)
        return out
    dgeom = {"rings": dist[0]["geometry"]["coordinates"]} \
        if dist[0]["geometry"]["type"] == "Polygon" else \
        {"rings": [r for poly in dist[0]["geometry"]["coordinates"] for r in poly]}

    own = field_like(238, "OWN", "DESC") or field_like(238, "OWN")
    if not own:
        out["error"] = "no ownership field on layer 238"
        return out
    out["ownership_field"] = own
    values = distinct(238, own)
    out["ownership_values"] = values
    private = [v for v in values if any(w in v.upper() for w in PRIVATE_WORDS)]
    out["private_values"] = private
    if not private:
        out["error"] = "no private ownership class found among: " + \
            ", ".join(values[:12])
        return out

    where = "{} IN ({})".format(own, ",".join("'{}'".format(q(v))
                                              for v in private))
    parcels = geometry_of(238, where, limit=parcels)
    out["private_parcels_sampled"] = len(parcels)
    if not parcels:
        out["error"] = "no private ownership polygons returned"
        return out

    hits, checked = [], 0
    yr = ""
    for f in schema(543):
        if "HARVEST" in f["name"].upper() and "YEAR" in f["name"].upper():
            yr = f["name"]
            break
    recent_where = "{} >= {}".format(yr, out.get("since_year", 2020)) if yr else "1=1"
    for p in parcels:
        g = p.get("geometry") or {}
        if g.get("type") == "Polygon":
            ring = {"rings": g["coordinates"]}
        elif g.get("type") == "MultiPolygon":
            ring = {"rings": [r for poly in g["coordinates"] for r in poly]}
        else:
            continue
        checked += 1
        n = count_within(543, ring)
        if n > 0:
            fields = ",".join(x for x in ("OPENING_ID", yr, "AREA_HA") if x)
            sample = sample_within(543, ring, fields)
            # Recency is what decides whether this is useful. A 1986 harvest
            # says the layer once saw private land; a 2024 one says it still
            # does, which is the only version that helps a DDS.
            recent = count_within(543, ring, recent_where)
            # A negative OPENING_ID means the polygon did not come from
            # RESULTS - it came from the satellite change detection feed,
            # which is precisely why it can see private land at all.
            detected = sum(1 for r in sample
                           if isinstance(r.get("OPENING_ID"), (int, float))
                           and r["OPENING_ID"] < 0)
            hits.append({"harvests": n, "recent": recent,
                         "from_change_detection": detected,
                         "sample": sample[:3]})
        time.sleep(0.3)

    out["parcels_checked"] = checked
    out["parcels_with_harvest"] = len(hits)
    out["hits"] = hits
    out["recent_harvests"] = sum(h.get("recent", 0) for h in hits)
    out["since_year_used"] = out.get("since_year", 2020)

    if not hits:
        out["verdict"] = ("no harvests found on the private parcels sampled - "
                          "consistent with the documentation that 543 is "
                          "crown land only")
    elif out["recent_harvests"]:
        out["verdict"] = (
            "543 holds RECENT harvests on private land ({} since {}). This is "
            "the first public source of observed harvest polygons for class B "
            "- unattributed, but real cuts with a year on them.".format(
                out["recent_harvests"], out["since_year_used"]))
    else:
        out["verdict"] = (
            "543 holds harvests on private land, but none recent in this "
            "sample. Historic coverage only is of little use for a current "
            "DDS - widen the sample before concluding either way.")
    return out


def test_814_overlap(path: str) -> dict:
    """Do any of our resolved harvest areas fall inside a Fibre Recovery Zone?

    The zones are small - single hectares - so an overlap would mean something
    quite specific rather than a general catchment. Worth knowing either way.
    """
    out: dict = {"file": path}
    try:
        with open(path, encoding="utf-8") as fh:
            gj = json.load(fh)
    except Exception as exc:
        out["error"] = str(exc)
        return out
    feats = gj.get("features") or []
    out["features_in_file"] = len(feats)
    box = bbox_polygon(feats)
    if not box:
        out["error"] = "no coordinates found in that file"
        return out

    out["zones_in_bbox"] = count_within(814, box)
    if out["zones_in_bbox"] > 0:
        tfield = field_like(814, "ZONE", "TYPE") or "FIBRE_RECOVERY_ZONE_TYPE"
        out["zone_types"] = distinct(814, tfield)[:10]
        out["sample"] = sample_within(814, box,
                                      ",".join([tfield, "FEATURE_AREA_SQM"]))
    per = []
    for f in feats[:40]:
        g = f.get("geometry") or {}
        if g.get("type") == "Polygon":
            ring = {"rings": g["coordinates"]}
        elif g.get("type") == "MultiPolygon":
            ring = {"rings": [r for poly in g["coordinates"] for r in poly]}
        else:
            continue
        n = count_within(814, ring)
        if n > 0:
            per.append({"identifier": (f.get("properties") or {}).get(
                "harp_identifier"), "zones": n})
        time.sleep(0.2)
    out["features_checked"] = min(40, len(feats))
    out["features_touching_a_zone"] = len(per)
    out["matches"] = per
    out["verdict"] = ("some resolved areas fall inside a Fibre Recovery Zone"
                      if per else
                      "no overlap - the zones are unrelated to these harvests")
    return out


def probe_layer(layer: int, sample: int = 5) -> dict:
    """What a layer is, whether it is queryable, and what it looks like."""
    out = {"layer": layer, "name": LAYERS.get(layer, "?")}
    url = "{}/{}".format(ROOT, layer)
    try:
        meta = post(url, {"f": "json"})
        out["service_name"] = meta.get("name", "")
        out["geometry_type"] = meta.get("geometryType", "")
        out["description"] = re.sub(r"\s+", " ",
                                    str(meta.get("description") or ""))[:300]
        out["fields"] = [f["name"] for f in meta.get("fields", [])
                         if f.get("type") != "esriFieldTypeGeometry"]
    except Exception as exc:
        out["error"] = str(exc)[:120]
        return out

    try:
        out["records"] = post(url + "/query",
                              {"where": "1=1", "returnCountOnly": "true",
                               "f": "json"}).get("count", 0)
    except Exception:
        out["records"] = "?"

    try:
        fields = ",".join(out["fields"][:12]) or "*"
        data = post(url + "/query", {"where": "1=1", "outFields": fields,
                                     "returnGeometry": "false",
                                     "resultRecordCount": sample, "f": "json"})
        out["sample"] = [f["attributes"] for f in data.get("features", [])]
    except Exception as exc:
        out["sample"] = []
        out["sample_error"] = str(exc)[:90]
    return out


# ──────────────────────────────── input ────────────────────────────────────

def suppliers_from_register(path: str, klass: str | None = "C2") -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("--register needs openpyxl:  pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)

    # Try every sheet rather than two hardcoded names. The register has been
    # through several shapes - source-level with a 'Register' sheet, then
    # supplier-level with a 'Suppliers' sheet - and a reader that knows only
    # last month's layout fails on a workbook that is perfectly readable.
    preferred = ("Suppliers", "Register", "Sources_Detail", "All_suppliers",
                 "Catchments_needed")
    order = [s for s in preferred if s in wb.sheetnames]
    order += [s for s in wb.sheetnames if s not in order]

    tried = []
    for name in order:
        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            continue
        # a title row may sit above the header, so look a few rows down
        for header_row in range(0, min(6, len(rows))):
            hdr = [str(h or "").strip() for h in rows[header_row]]
            low = [h.lower() for h in hdr]
            i_sup = next((low.index(n) for n in
                          ("supplier name", "supplier", "name")
                          if n in low), None)
            if i_sup is None:
                continue
            cands = [i for i, h in enumerate(low)
                     if re.fullmatch(r"class\s*v?\d*", h) or h == "class"]
            i_cls = sorted(cands, reverse=True)[0] if cands else None
            i_jur = next((low.index(n) for n in
                          ("province / state", "jurisdiction", "stateid")
                          if n in low), None)
            rows = rows[header_row:]
            break
        else:
            tried.append("{} ({})".format(
                name, ", ".join(h for h in
                                [str(x or "").strip() for x in rows[0]] if h)[:60]))
            continue
        break
    else:
        sys.exit("No supplier column found in {}.\n\nSheets tried:\n  {}"
                 .format(os.path.basename(path), "\n  ".join(tried) or "none"))

    seen = []
    for r in rows[1:]:
        if i_cls is not None and klass and i_cls < len(r):
            # A supplier-level row can carry several classes at once, so a
            # contains test rather than equality.
            cell = str(r[i_cls] or "").strip().upper()
            if klass.upper() not in [c.strip() for c in cell.split(",")]:
                continue
        if i_jur is not None and i_jur < len(r):
            jur = str(r[i_jur] or "").upper()
            if jur and "BC" not in jur and "BRITISH" not in jur:
                continue      # only BC has FTEN
        v = str(r[i_sup] or "").strip()
        if v and v not in seen:
            seen.append(v)
    return seen


# ──────────────────────────────── main ─────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--suppliers", help="comma separated supplier names")
    ap.add_argument("--register", help="a HARP register xlsx to read them from")
    ap.add_argument("--class", dest="klass", default="C2",
                    help="class filter for --register (default C2)")
    ap.add_argument("--district", default="", help="narrow to one district code, "
                                                   "for every supplier")
    ap.add_argument("--aliases", metavar="CSV",
                    default="./data/registry/supplier_aliases.csv",
                    help="the supplier alias table. Accepted rows are used "
                         "without searching; new candidates are proposed into "
                         "it for a person to rule on")
    ap.add_argument("--no-aliases", action="store_true",
                    help="ignore the table and re-derive everything")
    ap.add_argument("--districts-from", metavar="CSV",
                    help="a supplier_locations csv from `harp mills`. "
                         "Each supplier is narrowed to its own district, which "
                         "is the only way this measures anything - most of "
                         "these suppliers are not in the same district")
    ap.add_argument("--since", default="", help="harvest start after YYYY-MM-DD")
    ap.add_argument("--layers-only", action="store_true")
    ap.add_argument("--probe-layers", action="store_true",
                    help="re-run the layer and HBS probes settled on 12 Aug")
    ap.add_argument("--test-543", action="store_true",
                    help="does layer 543 hold harvests on private land?")
    ap.add_argument("--test-814", metavar="GEOJSON",
                    help="do any resolved areas fall inside a Fibre Recovery "
                         "Zone? pass a HARP areas geojson")
    ap.add_argument("--test-district", default="DSI",
                    help="district for --test-543 (default DSI)")
    ap.add_argument("--since-year", type=int, default=2020,
                    help="what counts as a recent harvest (default 2020)")
    ap.add_argument("--parcels", type=int, default=8,
                    help="private parcels to sample for --test-543")
    ap.add_argument("--skip-hbs", action="store_true")
    ap.add_argument("--out", default="c2_probe_output")
    ap.add_argument("--limit", type=int)
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    print(ATTRIBUTION)

    # The layer and HBS probes were settled on 12 Aug. Kept, because a settled
    # finding should stay checkable, but skipped unless asked for - they are
    # noise on a run whose purpose is the supplier envelopes.
    show_probes = args.probe_layers or args.layers_only or not (
        args.suppliers or args.register)

    # ---- 3. layers ----
    if show_probes:
        print("\n" + "=" * 78)
        print("3.  UNUSED LAYERS")
        print("=" * 78)
        layer_rows = []
        for layer in (793, 814, 543):
            info = probe_layer(layer)
            layer_rows.append(info)
            print("\n{}  {}".format(layer, info.get("service_name") or info["name"]))
            if info.get("error"):
                print("    unreachable: {}".format(info["error"]))
                continue
            print("    geometry : {}".format(info.get("geometry_type")))
            print("    records  : {:,}".format(info["records"])
                  if isinstance(info["records"], int) else "    records  : ?")
            if info.get("description"):
                print("    about    : {}".format(info["description"][:150]))
            print("    fields   : {}".format(", ".join(info["fields"][:10])))
            for s in (info.get("sample") or [])[:2]:
                vals = {k: v for k, v in list(s.items())[:6] if v not in (None, "")}
                print("    sample   : {}".format(vals))

        with open(os.path.join(args.out, "layers_{}.json".format(stamp)), "w",
                  encoding="utf-8") as fh:
            json.dump(layer_rows, fh, indent=1)

        # ---- 3b. does 543 cover private land? ----
        if args.test_543:
            print("\n" + "=" * 78)
            print("3b. LAYER 543 - does it hold harvests on PRIVATE land?")
            print("=" * 78)
            r = test_543_private(args.test_district, args.since_year,
                                 args.parcels)
            if r.get("error"):
                print("  could not test: {}".format(r["error"]))
            else:
                print("  ownership field  : {}".format(r["ownership_field"]))
                print("  private classes  : {}".format(
                    ", ".join(r["private_values"])[:120]))
                print("  parcels checked  : {}".format(r["parcels_checked"]))
                print("  with a harvest   : {}".format(r["parcels_with_harvest"]))
                print("  harvests since {} : {}".format(
                    r["since_year_used"], r["recent_harvests"]))
                for h in r["hits"][:6]:
                    print("      {} harvest(s), {} since {}, {} from change "
                          "detection".format(h["harvests"], h.get("recent", 0),
                                             r["since_year_used"],
                                             h.get("from_change_detection", 0)))
                    for x in h["sample"][:2]:
                        print("         {}".format(x))
            print("\n  -> {}".format(r.get("verdict", "inconclusive")))
            with open(os.path.join(args.out, "test543_{}.json".format(stamp)),
                      "w", encoding="utf-8") as fh:
                json.dump(r, fh, indent=1)

        # ---- 3c. do our areas touch a Fibre Recovery Zone? ----
        if args.test_814:
            print("\n" + "=" * 78)
            print("3c. LAYER 814 - do resolved areas fall inside a Fibre Recovery Zone?")
            print("=" * 78)
            r = test_814_overlap(args.test_814)
            if r.get("error"):
                print("  could not test: {}".format(r["error"]))
            else:
                print("  features in file      : {}".format(r["features_in_file"]))
                print("  zones in bounding box : {}".format(r["zones_in_bbox"]))
                if r.get("zone_types"):
                    print("  zone types            : {}".format(
                        ", ".join(r["zone_types"])[:110]))
                print("  features checked      : {}".format(r["features_checked"]))
                print("  touching a zone       : {}".format(
                    r["features_touching_a_zone"]))
                for m in r["matches"][:8]:
                    print("      {} -> {} zone(s)".format(m["identifier"], m["zones"]))
            print("\n  -> {}".format(r.get("verdict", "inconclusive")))
            with open(os.path.join(args.out, "test814_{}.json".format(stamp)),
                      "w", encoding="utf-8") as fh:
                json.dump(r, fh, indent=1)

        if args.layers_only or args.test_543 or args.test_814:
            print("\nWritten to {}".format(args.out))
            if not (args.suppliers or args.register):
                return

        # ---- 2. HBS by client ----
        if not args.skip_hbs:
            print("\n" + "=" * 78)
            print("2.  HBS BY CLIENT - is anything public keyed on client number?")
            print("=" * 78)
            rows = probe_hbs()
            for r in rows:
                flag = ""
                if r.get("looks_like_login"):
                    flag = "  <- login wall"
                elif r.get("status") == 200 and r.get("has_client_field"):
                    flag = "  <- mentions client, worth a look"
                print("  {:<34} {:>6}  {:>7}b  {}{}".format(
                    r["probe"][:34], str(r.get("status")), r.get("bytes", 0),
                    (r.get("title") or "")[:34], flag))
            with open(os.path.join(args.out, "hbs_probe_{}.csv".format(stamp)),
                      "w", encoding="utf-8-sig", newline="") as fh:
                w = csv.DictWriter(fh, fieldnames=sorted(
                    {k for r in rows for k in r}))
                w.writeheader()
                w.writerows(rows)
            print("\n  SETTLED: tested in a browser 12 Aug 2026. The client-keyed"
                  " URL renders the\n  mark query form and treats clientNumber as"
                  " the mark - truncated to six\n  characters and rejected. There"
                  " is no public client search.")

    # ---- 1. operator envelope ----
    names = []
    if args.suppliers:
        names = [n.strip() for n in args.suppliers.split(",") if n.strip()]
    elif args.register:
        names = suppliers_from_register(args.register, args.klass)
    if args.limit:
        names = names[:args.limit]

    if not names:
        print("\nNo suppliers given - skipping the envelope test. "
              "Use --suppliers or --register.")
        return

    print("\n" + "=" * 78)
    print("1.  OPERATOR ENVELOPE - {} suppliers".format(len(names)))
    # Per-supplier districts. A single district for everyone answers the wrong
    # question: Ludwig is Campbell River, Waldun is Chilliwack, Aspen Planers
    # is Thompson Rivers. Running all of them against South Island returned
    # zeros that looked like a failed route rather than a bad filter.
    per_district = {}
    if args.districts_from:
        import csv as _csv
        try:
            with open(args.districts_from, encoding="utf-8-sig", newline="") as fh:
                for row in _csv.DictReader(fh):
                    code = (row.get("district_code") or "").strip()
                    name = (row.get("supplier") or "").strip()
                    if code and name:
                        per_district[name.upper()] = code
        except Exception as exc:
            sys.exit("Could not read {}: {}".format(args.districts_from, exc))
        print("\n{} supplier(s) have a district of their own".format(
            len(per_district)))
        placed = sum(1 for n in names if n.upper() in per_district)
        print("{} of {} on this run will be narrowed; the rest run "
              "province-wide and are marked as such".format(placed, len(names)))

    if args.district or args.since or per_district:
        bits = []
        if per_district:
            bits.append("each supplier's own district")
        elif args.district:
            bits.append("district " + args.district)
        if args.since:
            bits.append("harvest start after " + args.since)
        print("    narrowed by: " + ", ".join(bits))
    print("=" * 78)
    print("\n{:<32} {:<7} {:<34} {:>6} {:>8} {:>8} {:>7}".format(
        "SUPPLIER", "MATCH", "FTEN CLIENT", "DIST", "ALL", "IN DIST",
        "WINDOW"))
    print("-" * 110)

    # The alias table decides; the matcher only proposes. A confirmation made
    # last month holds this month, and tightening the matcher cannot quietly
    # change a historical answer.
    table = None
    if not args.no_aliases:
        sys.path.insert(0, os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))))
        from harp.aliases import AliasTable
        table = AliasTable(args.aliases)
        if len(table.rows):
            print("\nalias table: " + table.summary().splitlines()[0])

    rows = []
    new_proposals = 0
    for name in names:
        accepted = table.for_supplier(name) if table else []
        if accepted:
            tier = "accepted (in the alias table)"
            clients = [{"client_name": a.client_name,
                        "client_number": a.client_number,
                        "location": "", "verified": "accepted"}
                       for a in accepted]
        else:
            tier, clients = find_clients(name)
            if table:
                # Anything the matcher is confident of is proposed, not used.
                # Only an exact name is taken automatically.
                for c in clients:
                    if table.auto_accept(c.get("verified", "")):
                        table.decide(name, c["client_number"], True,
                                     who="matcher",
                                     reason="exact name match")
                    else:
                        if table.propose(name, c["client_number"],
                                         c["client_name"],
                                         basis=c.get("verified", "")) == "proposed":
                            new_proposals += 1
                # Plausible-but-unconfirmable matches go in the table too.
                # They are exactly what a person is needed for, and leaving
                # them only in console output means they are re-derived and
                # re-forgotten every run.
                for cand in POSSIBLE.get(name, []):
                    if table.propose(name, cand["client_number"],
                                     cand["client_name"],
                                     basis="possible - shares the name but "
                                           "carries a word of its own"
                                     ) == "proposed":
                        new_proposals += 1
        if not clients:
            print("{:<34} {:<9} {}".format(name[:34], "none",
                                           "no FTEN tenure under this name"))
            rows.append({"supplier": name, "match_tier": "", "client_name": "",
                         "client_number": "", "location": "", "blocks_all": 0,
                         "blocks_district": "", "blocks_window": "",
                         "area_ha": 0, "note": "no tenure found"})
            continue

        short = "high" if tier.startswith("high") else (
            "medium" if tier.startswith("medium") else "low")
        for c in clients[:6]:
            district_here = per_district.get(name.upper(), args.district)
            env = envelope(c["client_number"], c["location"], district_here,
                           args.since)
            env["district_used"] = district_here or "(province-wide)"
            print("{:<32} {:<7} {:<34} {:>6} {:>8} {:>8} {:>7}".format(
                name[:32], short, c["client_name"][:34],
                district_here or "—", env["blocks_all"],
                env.get("blocks_district", "-"),
                env.get("blocks_window", "-")))
            rows.append({"supplier": name, "match_tier": tier,
                         "client_name": c["client_name"],
                         "client_number": c["client_number"],
                         "location": c["location"],
                         "blocks_all": env["blocks_all"],
                         "blocks_district": env.get("blocks_district", ""),
                         "blocks_window": env.get("blocks_window", ""),
                         "area_ha": env.get("area_ha", 0),
                         "area_truncated": env.get("area_truncated", False),
                         "district_used": env.get("district_used", ""),
                         "where": env.get("where", ""),
                         "note": ""})
            time.sleep(0.2)
        if len(clients) > 6:
            print("{:<34} {:<9} ... and {} more candidates".format(
                "", "", len(clients) - 6))

    path = os.path.join(args.out, "envelopes_{}.csv".format(stamp))
    cols = ["supplier", "match_tier", "client_name", "client_number",
            "location", "district_used", "blocks_all", "blocks_district",
            "blocks_window", "area_ha", "area_truncated", "where", "note"]
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})

    found = sum(1 for r in rows if r["blocks_all"])
    high = sum(1 for r in rows if str(r["match_tier"]).startswith("high"))
    print("\n" + "-" * 78)
    print("  {} of {} rows returned tenure, {} on a high-confidence name match"
          .format(found, len(rows), high))
    if args.district or per_district:
        # Only rows that actually had a district applied. Counting the
        # province-wide ones as though they had been narrowed would overstate
        # the filter.
        narrowed = [r for r in rows
                    if r.get("district_used") and
                    r["district_used"] != "(province-wide)"]
        tot = sum(int(r["blocks_district"] or 0) for r in narrowed
                  if str(r["blocks_district"]).isdigit())
        allb = sum(int(r["blocks_all"] or 0) for r in narrowed)
        if allb:
            print("  of the {} row(s) with a district: {:,} blocks cut to "
                  "{:,} ({:.0f}% narrower)".format(
                      len(narrowed), allb, tot, 100 - tot / allb * 100))
        skipped = len(rows) - len(narrowed)
        if skipped:
            print("  {} row(s) ran province-wide — no district known for that "
                  "supplier".format(skipped))
        print("  A district envelope is everywhere that operator cut in that "
              "district, not what they sold Harmac. Detection narrows it "
              "further; it is P2 and never a plot claim.")
    else:
        print("  Run again with --district and --since. An envelope covering a "
              "supplier's whole provincial tenure is not useful; the same "
              "envelope cut to one district and one delivery year might be.")
    if table:
        table.save()
        print("\n  alias table: {}".format(table.path))
        for line in table.summary().splitlines():
            print("    " + line)
        waiting = len(table.open_questions())
        if waiting:
            print("    {} proposal(s) are NOT being used until somebody rules "
                  "on them.".format(waiting))
            print("    python tools/aliases.py review")

    if POSSIBLE:
        print("\n  {} supplier(s) had a plausible but unconfirmable match. "
              "These are NOT included above:".format(len(POSSIBLE)))
        for sup, cands in list(POSSIBLE.items())[:14]:
            print("    {:<30} {}".format(
                sup[:30], ", ".join(c["client_name"] for c in cands[:3])[:70]))
        print("    Each shares the supplier's name but carries a word of its "
              "own. Someone has to say whether they are the same company.")
    print("  Written to {}".format(path))
    print("-" * 78)


if __name__ == "__main__":
    main()

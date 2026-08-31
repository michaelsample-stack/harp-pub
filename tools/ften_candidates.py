#!/usr/bin/env python3
"""Find FTEN tenure candidates for the suppliers we have not placed.

A different job from the matcher in c2_probe. That one decides, and decides
conservatively, because a wrong holder is thousands of blocks of someone
else's forest. This one searches hard and ranks, because the person reading
the output is the filter.

    python tools/ften_candidates.py --suppliers "C&C Lath Mill Ltd, Aquila Cedar"
    python tools/ften_candidates.py --register HPA1_Supplier_Register_v7_0.xlsx
    python tools/ften_candidates.py --register ... --min-score 40

WHAT IT DOES DIFFERENTLY
------------------------
Searches several ways rather than one: the full name, each identifying word,
each pair of words, and known expansions of forestry abbreviations. Then
scores every candidate on how much of the supplier's name it accounts for and
how much of its own name is unexplained.

Nothing is accepted or rejected. Everything above the score floor is listed
for a person, and the accepted ones go into the alias table where the decision
persists.

WHY NOT JUST LOOSEN THE MATCHER
-------------------------------
Because then the loosened rule runs unattended next month and quietly attaches
Dunkley Lumber to Ludwig. Searching wide is safe as long as a person is
reading; it is not safe as a default.

WHAT IT CANNOT DO
-----------------
Some supplier codes have no company name anywhere in Harmac's data - NVCL is
recorded as "NVCL NVCL", and COS, WEW, WWK and WWW carry no expansion at all.
There is nothing to search for, and no amount of querying substitutes for
asking the client what the code means.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from datetime import datetime

import requests

BASE = ("https://delivery.maps.gov.bc.ca/arcgis/rest/services"
        "/mpcm/bcgwpub/MapServer")
BLOCKS = "{}/340/query".format(BASE)
TIMEOUT = 120

S = requests.Session()
S.headers.update({"User-Agent": "NGIS-HARP-candidates/1.0"})

# Words shared across the industry. Present in a name, they say nothing about
# which company it is.
GENERIC = {
    "LTD", "LTD.", "INC", "INC.", "LIMITED", "CORP", "CORPORATION", "COMPANY",
    "CO", "CO.", "LP", "LLP", "GP", "THE", "AND", "GROUP", "HOLDINGS",
    "PARTNERSHIP", "ENTERPRISES", "INDUSTRIES", "PRODUCTS", "FOREST",
    "FORESTRY", "FORESTS", "LOGGING", "LUMBER", "TIMBER", "WOOD", "WOODS",
    "FIBRE", "FIBER", "MILL", "MILLS", "SAWMILL", "SAWMILLS", "SHAKE",
    "SHAKES", "SHINGLE", "SHINGLES", "CEDAR", "FIR", "HEMLOCK", "SPRUCE",
    "PINE", "CONTRACTING", "SERVICES", "MANAGEMENT", "RESOURCE", "RESOURCES",
    "TRADING", "SUPPLY", "CHIPPING", "CUSTOM", "SPLIT", "BARGE", "PULP",
    "PAPER", "PLYWOOD", "VENEER", "CHIP", "CHIPS", "LOG", "LOGS", "VALUE",
    "ADDED", "RELOAD", "CANADA", "CANADIAN", "BC", "BRITISH", "COLUMBIA",
}

# Expansions worth trying. Harmac writes a short code; FTEN carries the full
# registered name, and no string comparison bridges the two.
EXPANSIONS = {
    "WFP": ["WESTERN FOREST PRODUCTS"],
    "S&R": ["S & R", "S AND R"],
    "C&C": ["C & C", "C AND C"],
    "G&R": ["G & R", "G AND R"],
    "SSD": ["SSD SAWMILL SALES DIRECT"],
    "NVCL": [],          # no expansion known - ask the client
    "COS": [], "WEW": [], "WWK": [], "WWW": [],
}


def q(v) -> str:
    return str(v).replace("'", "''")


def canon(name: str) -> str:
    n = re.sub(r"\(.*?\)", " ", str(name or "")).upper()
    n = n.replace("&", " AND ")
    n = re.sub(r"[^A-Z0-9\s'-]", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def words_of(name: str) -> list[str]:
    """The words that identify a company.

    Short initials count when they are all a name has. 'S&R Sawmills' reduces
    to nothing under the usual rules - S and R are too short, SAWMILLS is
    generic - and a name that reduces to nothing can never be matched. For
    those, the initials are the identity.
    """
    toks = [w for w in canon(name).split() if w != "AND"]
    kept = [w for w in toks if w not in GENERIC and len(w) >= 3]
    if kept:
        return kept
    initials = [w for w in toks if w not in GENERIC and len(w) <= 2]
    return initials or [w for w in toks if w not in GENERIC][:2]


def search(term: str) -> list[dict]:
    """Every FTEN client whose name contains this string."""
    try:
        r = S.post(BLOCKS, data={
            "where": "CLIENT_NAME LIKE '%{}%'".format(q(term).upper()),
            "outFields": "CLIENT_NAME,CLIENT_NUMBER,CLIENT_LOCATION_CODE",
            "returnDistinctValues": "true", "returnGeometry": "false",
            "resultRecordCount": 300, "f": "json"}, timeout=TIMEOUT)
        r.raise_for_status()
        data = r.json()
    except Exception:
        return []
    out = []
    for f in data.get("features", []):
        a = f["attributes"]
        if a.get("CLIENT_NAME"):
            out.append({"client_name": a["CLIENT_NAME"].strip(),
                        "client_number": (a.get("CLIENT_NUMBER") or "").strip(),
                        "location": (a.get("CLIENT_LOCATION_CODE") or "").strip()})
    return out


def block_count(client_number: str) -> int:
    try:
        r = S.post(BLOCKS, data={
            "where": "CLIENT_NUMBER = '{}'".format(q(client_number)),
            "returnCountOnly": "true", "f": "json"}, timeout=TIMEOUT)
        r.raise_for_status()
        return int(r.json().get("count") or 0)
    except Exception:
        return 0


def profile(client_number: str, sample: int = 400) -> dict:
    """Where this client actually operates, and what their marks look like.

    A score compares two strings. This is the evidence that decides whether
    the company behind the string is the right one: which districts they cut
    in, how much ground, and a handful of real timber marks to eyeball.

    Sampled rather than exhaustive - a holder with eleven thousand blocks
    would take minutes to enumerate and the shape of the first few hundred
    tells you what you need.
    """
    out = {"districts": {}, "marks": [], "area_ha": 0.0, "years": [],
           "locations": set()}
    try:
        r = S.post(BLOCKS, data={
            "where": "CLIENT_NUMBER = '{}'".format(q(client_number)),
            "outFields": ("TIMBER_MARK,GEOGRAPHIC_DISTRICT_CODE,"
                          "GEOGRAPHIC_DISTRICT_NAME,FEATURE_AREA,"
                          "DISTURBANCE_START_DATE,CLIENT_LOCATION_CODE,"
                          "CLIENT_NAME"),
            "returnGeometry": "false", "resultRecordCount": sample,
            "f": "json"}, timeout=TIMEOUT)
        r.raise_for_status()
        feats = r.json().get("features", [])
    except Exception:
        return out

    for f in feats:
        a = f.get("attributes", {})
        d = a.get("GEOGRAPHIC_DISTRICT_CODE")
        if d:
            out["districts"][d] = out["districts"].get(d, 0) + 1
        m = a.get("TIMBER_MARK")
        if m and m not in out["marks"]:
            out["marks"].append(m)
        try:
            out["area_ha"] += float(a.get("FEATURE_AREA") or 0) / 10000.0
        except (TypeError, ValueError):
            pass
        lc = a.get("CLIENT_LOCATION_CODE")
        if lc:
            out["locations"].add(str(lc))
        ts = a.get("DISTURBANCE_START_DATE")
        if ts:
            try:
                out["years"].append(datetime.utcfromtimestamp(
                    int(ts) / 1000).year)
            except Exception:
                pass
    out["sampled"] = len(feats)
    out["locations"] = sorted(out["locations"])
    return out


def score(supplier: str, client: str) -> tuple[int, str]:
    """How well does this client account for the supplier's name?

    Two halves, because both matter and they fail differently. Coverage is how
    much of the supplier's name the client carries - low coverage means we
    found a different company. Cleanliness is how much of the CLIENT's name is
    unexplained - Imperial Oil covers 'Imperial' completely and is still the
    wrong firm, because OIL is unaccounted for.
    """
    sw, cw = words_of(supplier), words_of(client)
    if not sw:
        return 0, "supplier name is all generic words"
    cset = set(cw)

    def near(w, pool):
        return w in pool or any(
            abs(len(c) - len(w)) <= 2 and (c.startswith(w) or w.startswith(c))
            for c in pool)

    covered = [w for w in sw if near(w, cset)]
    coverage = len(covered) / len(sw)
    unexplained = [w for w in cw if not near(w, set(sw))]
    cleanliness = 1 - (len(unexplained) / len(cw)) if cw else 1

    s = int(round(100 * (0.65 * coverage + 0.35 * cleanliness)))
    if canon(supplier) == canon(client):
        s = 100

    bits = []
    if coverage == 1:
        bits.append("every supplier word present")
    else:
        bits.append("{} of {} supplier words".format(len(covered), len(sw)))
    if unexplained:
        bits.append("client also says " + ", ".join(unexplained[:3]))
    return s, "; ".join(bits)


def candidates(supplier: str, log=print) -> list[dict]:
    """Search several ways, then rank what comes back."""
    code = supplier.strip().upper()
    terms: list[str] = []
    full = canon(supplier)
    if full:
        terms.append(full)
    ws = words_of(supplier)
    terms.extend(ws[:5])
    for i in range(len(ws) - 1):
        terms.append("{} {}".format(ws[i], ws[i + 1]))
    for key, exp in EXPANSIONS.items():
        if key in code:
            terms.extend(exp)

    seen: dict[tuple, dict] = {}
    for t in dict.fromkeys(terms):
        if len(t) < 3:
            continue
        for c in search(t):
            key = (c["client_name"], c["client_number"])
            if key not in seen:
                c["found_by"] = t
                seen[key] = c
        time.sleep(0.1)

    out = []
    for c in seen.values():
        s, why = score(supplier, c["client_name"])
        c["score"], c["why"] = s, why
        out.append(c)
    out.sort(key=lambda c: -c["score"])
    return out


# A source identifier is usually the mill town - PARKSVILLE, DUNCAN,
# CASTLEGAR. It is not a harvest location, but it says which part of the
# province the company works in, and that is often what separates two
# candidates with the same score.
REGION_OF = {
    "DSI": "south island", "DCR": "campbell river", "DNI": "north island",
    "DSC": "sunshine coast", "DCK": "chilliwack", "DCC": "cascades",
    "DKA": "thompson", "DOS": "okanagan", "DKM": "skeena", "DND": "nadina",
    "DQU": "quesnel", "DPC": "prince george", "DMK": "mackenzie",
    "DSE": "selkirk", "DRM": "rocky mountain", "DKA": "thompson",
}
ISLAND = {"DSI", "DCR", "DNI"}
COAST = {"DSC", "DCK", "DCC"} | ISLAND

TOWN_REGION = {
    "PARKSVILLE": ISLAND, "DUNCAN": ISLAND, "NANAIMO": ISLAND,
    "CAMPBELL RIVER": ISLAND, "PORT ALBERNI": ISLAND, "PORTALBERNI": ISLAND,
    "GOLD RIVER": ISLAND, "CHEMAINUS": ISLAND, "ERRINGTON": ISLAND,
    "BLACKCREEK": ISLAND, "PORT MCNEILL": ISLAND, "DUKE": ISLAND,
    "COURTENAY": ISLAND, "LADYSMITH": ISLAND,
    "ABBOTSFORD": {"DCK"}, "MISSION": {"DCK"}, "MAPLE-RIDGE": {"DCK"},
    "MAPLE RIDGE": {"DCK"}, "PITT MEADOWS": {"DCK"}, "SURREY": {"DCK"},
    "DELTA": {"DCK"}, "RICHMOND": {"DCK"}, "SILVERDALE": {"DCK"},
    "CASTLEGAR": {"DSE", "DRM"}, "PRINCETON": {"DCC", "DOS"},
    "MERRITT": {"DCC"}, "KAMLOOPS": {"DKA"},
}


def load_hints(path: str) -> dict:
    """Mill town per supplier, from their source identifiers."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    hints: dict[str, set] = {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            rows = list(wb[sheet].iter_rows(values_only=True))
            if not rows:
                continue
            low = [str(h or "").strip().lower() for h in rows[0]]
            i_s = next((low.index(n) for n in ("supplier", "supplier name")
                        if n in low), None)
            i_k = next((low.index(n) for n in ("alias / key in that system",
                                               "alias_or_key", "source codes")
                        if n in low), None)
            if i_s is None or i_k is None:
                continue
            for r in rows[1:]:
                if i_s >= len(r) or not r[i_s]:
                    continue
                key = str(r[i_s]).strip().upper()
                text = str(r[i_k] or "").upper()
                for town, districts in TOWN_REGION.items():
                    if town in text:
                        hints.setdefault(key, set()).update(districts)
    except Exception:
        return {}
    return hints


def geography_check(hint: set, districts: dict) -> tuple[str, str]:
    """Does this candidate cut where the supplier's mill is?

    Not proof either way. A coastal mill can buy interior fibre and often
    does. But a candidate whose entire tenure is in the Kootenays, against a
    supplier milling in Parksville, is worth a second look before accepting.
    """
    if not hint or not districts:
        return "", ""
    total = sum(districts.values())
    inside = sum(n for d, n in districts.items() if d in hint)
    pct = inside / total * 100 if total else 0
    top = sorted(districts.items(), key=lambda kv: -kv[1])[:3]
    where = ", ".join("{} {}".format(d, n) for d, n in top)
    if pct >= 50:
        return "consistent", "{:.0f}% near the mill  ({})".format(pct, where)
    if pct > 0:
        return "partly", "{:.0f}% near the mill  ({})".format(pct, where)
    return "elsewhere", "none near the mill  ({})".format(where)


def suppliers_from_register(path: str, klass: str | None) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("--register needs openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    order = [s for s in ("Suppliers", "Catchments_needed", "Register")
             if s in wb.sheetnames] + list(wb.sheetnames)
    for sheet in dict.fromkeys(order):
        rows = list(wb[sheet].iter_rows(values_only=True))
        for hr in range(0, min(6, len(rows))):
            low = [str(h or "").strip().lower() for h in rows[hr]]
            i_sup = next((low.index(n) for n in ("supplier", "supplier name",
                                                 "name") if n in low), None)
            if i_sup is None:
                continue
            i_cls = next((i for i, h in enumerate(low)
                          if h == "class" or re.fullmatch(r"class\s*v?\d*", h)),
                         None)
            out = []
            for r in rows[hr + 1:]:
                if i_sup >= len(r) or not r[i_sup]:
                    continue
                if klass and i_cls is not None and i_cls < len(r):
                    cell = str(r[i_cls] or "").upper()
                    if klass.upper() not in [c.strip() for c in cell.split(",")]:
                        continue
                v = str(r[i_sup]).strip()
                if v and v not in out:
                    out.append(v)
            if out:
                return out
    sys.exit("no supplier column found")


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--suppliers", help="comma separated names")
    ap.add_argument("--register", help="a HARP register xlsx")
    ap.add_argument("--class", dest="klass", help="class filter for --register")
    ap.add_argument("--min-score", type=int, default=35,
                    help="floor for listing a candidate (default 35)")
    ap.add_argument("--top", type=int, default=6,
                    help="candidates shown per supplier (default 6)")
    ap.add_argument("--blocks", action="store_true",
                    help="kept for compatibility - block counts and profiles "
                         "are always fetched now, since a score without them "
                         "cannot be checked")
    ap.add_argument("--aliases",
                    default="./data/registry/supplier_aliases.csv",
                    help="skip suppliers already decided here")
    ap.add_argument("--out", default="candidate_out")
    args = ap.parse_args()

    if args.suppliers:
        names = [n.strip() for n in args.suppliers.split(",") if n.strip()]
    elif args.register:
        names = suppliers_from_register(args.register, args.klass)
    else:
        ap.error("give --suppliers or --register")

    decided = set()
    if os.path.isfile(args.aliases):
        sys.path.insert(0, os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))))
        try:
            from harp.aliases import AliasTable
            t = AliasTable(args.aliases)
            decided = {a.supplier.strip().upper() for a in t.rows.values()
                       if a.state in ("accepted", "rejected")}
            if decided:
                print("{} supplier(s) already decided in the alias table - "
                      "skipping them".format(len(decided)))
        except Exception:
            pass

    os.makedirs(args.out, exist_ok=True)
    print("Contains information licensed under the Open Government Licence - "
          "British Columbia.\n")

    rows, no_name, nothing = [], [], []
    hints = load_hints(args.register) if args.register else {}

    for name in names:
        if name.strip().upper() in decided:
            continue
        code = name.strip().upper()
        if any(k in code and not v for k, v in EXPANSIONS.items()):
            no_name.append(name)
            continue

        cands = [c for c in candidates(name) if c["score"] >= args.min_score]
        if not cands:
            nothing.append(name)
            continue

        hint = hints.get(name.strip().upper(), set())
        print("\n" + "=" * 78)
        print("{}{}".format(name, "    mill near: " + ", ".join(sorted(hint))
                            if hint else ""))
        print("=" * 78)

        for c in cands[:args.top]:
            blocks = block_count(c["client_number"])
            c["blocks"] = blocks
            prof = profile(c["client_number"]) if blocks else {}
            geo, geo_why = geography_check(hint, prof.get("districts", {}))
            c["geography"] = geo
            c["districts"] = ", ".join(
                "{} {}".format(d, n) for d, n in
                sorted(prof.get("districts", {}).items(),
                       key=lambda kv: -kv[1])[:4])
            c["sample_marks"] = ", ".join(prof.get("marks", [])[:5])
            c["area_ha_sampled"] = round(prof.get("area_ha", 0), 1)
            yrs = prof.get("years") or []
            c["years"] = "{}–{}".format(min(yrs), max(yrs)) if yrs else ""
            c["locations"] = ", ".join(prof.get("locations", [])[:6])

            flag = {"consistent": "OK", "partly": "??", "elsewhere": "!!"}.get(
                geo, "  ")
            print("\n  {} {:<46} {}".format(flag, c["client_name"][:46],
                                            c["client_number"]))
            print("     score {:<4} {}".format(c["score"], c["why"][:58]))
            if blocks:
                print("     {:,} blocks   {:,.0f} ha in a {} block sample   {}"
                      .format(blocks, c["area_ha_sampled"],
                              prof.get("sampled", 0), c["years"]))
                if c["districts"]:
                    print("     districts: {}".format(c["districts"]))
                if geo_why:
                    print("     geography: {}".format(geo_why))
                if c["sample_marks"]:
                    print("     marks:     {}".format(c["sample_marks"]))
                if len(prof.get("locations", [])) > 1:
                    print("     locations: {}".format(c["locations"]))
            else:
                print("     no blocks under this client number")
            rows.append({"supplier": name,
                         "mill_near": ", ".join(sorted(hint)),
                         **{k: c[k] for k in
                            ("client_name", "client_number", "location",
                             "score", "why", "found_by", "blocks",
                             "geography", "districts", "sample_marks",
                             "area_ha_sampled", "years", "locations")
                            if k in c}})
        time.sleep(0.1)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(args.out, "ften_candidates_{}.csv".format(stamp))
    if rows:
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
            w.writeheader(); w.writerows(rows)

    print("\n" + "-" * 78)
    print("{} candidate(s) for {} supplier(s)".format(
        len(rows), len({r["supplier"] for r in rows})))
    if nothing:
        print("\n{} supplier(s) returned nothing above the score floor:"
              .format(len(nothing)))
        for n in nothing:
            print("  {}".format(n))
        print("  Either they hold no BC tenure, or they hold it under a name "
              "that shares nothing with what Harmac calls them.")
    if no_name:
        print("\n{} supplier(s) have no company name to search:"
              .format(len(no_name)))
        for n in no_name:
            print("  {}".format(n))
        print("  Harmac's own record carries the code and nothing else - NVCL "
              "is stored as 'NVCL NVCL'. This needs the client, not a query.")
    if rows:
        print("\n  {}".format(path))
        print("\nNothing here is a decision. Accept the right ones into the "
              "alias table:")
        print('  python tools/aliases.py accept "{}" {} --who YOU --reason "..."'
              .format(rows[0]["supplier"], rows[0]["client_number"]))


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Private timber mark -> parcel -> geometry.

The BC Ministry of Forests scaled-timbermark extracts link a PRIVATE timber
mark to the land parcels it was scaled from. That closes the one gap nothing
public could close: a private harvest has a real mark and no cutblock
geometry, but it does have a parcel, and ParcelMap BC publishes parcel
polygons under the Open Government Licence.

    timber mark  ->  PID  ->  ParcelMap BC  ->  polygon

WHAT THIS IS AND IS NOT
-----------------------
A parcel is the titled land the timber was scaled from. It is not the cut
block. A 400 ha parcel with a 30 ha cut on it gives you the 400 ha, and the
harvest is somewhere inside it. That is a real bounded plot of land - far
better than a district - but it must not be presented as a harvest boundary.

Where a mark has many parcels the total can be large: in the Harmac set, mark
'F' has 310 parcels and 'C23'' has 297. Read the area figures before deciding
what any of it can support.

THE FILES ARE NOT CUMULATIVE
----------------------------
Each monthly extract is a largely distinct set of marks, with pairwise overlap
of roughly 15-50%. Taking only the newest would discard most of them. This
builds the union of every file given and deduplicates on (mark, PID).

Filenames also do not match their contents - one workbook named "June 2026"
was processed in February and its data sheet is called "January 2026". So the
first sheet is read whatever it is called, and no period is inferred from a
filename.

Usage
-----
    python tools/ptm_parcels.py "BC_Private_Timber_Marks/"
    python tools/ptm_parcels.py ptm/ --marks EDRWD,AA545 --out parcels
    python tools/ptm_parcels.py ptm/ --register ..._v5.xlsx --class B
    python tools/ptm_parcels.py ptm/ --inventory-only

Requires: requests, openpyxl
Licence : Contains information licensed under the Open Government Licence -
          British Columbia. The source workbooks carry their own Legal
          Disclaimer sheet - read it before publishing derived geometry.
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import re
import sys
import time
from collections import Counter
from datetime import datetime

import requests

ROOT = "https://delivery.maps.gov.bc.ca/arcgis/rest/services/mpcm/bcgwpub/MapServer"
PARCELS = ROOT + "/218"
PAGE = 1000
TIMEOUT = 120
ATTRIBUTION = ("Contains information licensed under the "
               "Open Government Licence - British Columbia.")

WANTED = ("ORG_UNIT_CODE", "TIMBER_MARK", "PID", "NOTE", "LEGAL", "Source",
          "ProcessedOn")

S = requests.Session()
S.headers.update({"User-Agent": "NGIS-HARP-ptm/1.0"})


def post(url: str, params: dict) -> dict:
    r = S.post(url, data=params, timeout=TIMEOUT)
    r.raise_for_status()
    data = r.json()
    if "error" in data:
        raise RuntimeError(data["error"].get("message", str(data["error"])))
    return data


# ────────────────────────────── PID parsing ────────────────────────────────
#
# A PID is nine digits, usually written 000-189-065. The extracts also carry
# two other things in the same column:
#
#   comma lists with abbreviated continuations
#       009-664-726,-793,-823   ->  009664726, 009664793, 009664823
#       the later entries repeat the leading block of the first
#
#   land title document and plan numbers, which are not PIDs at all
#       CA7321377,78,99   CB1080949-53   FB522617,39-41,3104,7-9
#
# Six of 1,886 rows in the Harmac set are the second kind. They are flagged
# rather than dropped silently - a PID that cannot be parsed is a question for
# the client, not a rounding error.

PID_RE = re.compile(r"^\d{3}-?\d{3}-?\d{3}$")
TITLE_RE = re.compile(r"^[A-Z]{2}\d")


def parse_pids(raw: str) -> tuple[list[str], str]:
    """Every PID in a cell, plus a note on anything that is not one."""
    text = str(raw or "").strip()
    if not text:
        return [], "empty"
    if TITLE_RE.match(text):
        return [], "land title or plan number, not a PID: {}".format(text[:40])

    out, notes, prefix = [], [], ""
    for part in re.split(r"[,;]", text):
        p = part.strip()
        if not p:
            continue
        digits = re.sub(r"[^0-9]", "", p)
        if PID_RE.match(p) or len(digits) == 9:
            out.append(digits.zfill(9))
            prefix = digits.zfill(9)[:6]
            continue
        # an abbreviated continuation: -793 after 009-664-726
        if prefix and 1 <= len(digits) <= 3:
            out.append(prefix + digits.zfill(3))
            continue
        if digits:
            notes.append(p[:24])
    if not out and notes:
        return [], "unparsed: " + ", ".join(notes[:4])
    return out, ("partly unparsed: " + ", ".join(notes[:4])) if notes else ""


# ──────────────────────────────── loading ──────────────────────────────────

def load_extracts(path: str) -> tuple[list[dict], list[dict]]:
    """Union of every workbook in a folder. Returns (rows, per-file report)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("This needs openpyxl:  pip install openpyxl")

    files = sorted(glob.glob(os.path.join(path, "*.xlsx"))) \
        if os.path.isdir(path) else [path]
    if not files:
        sys.exit("No .xlsx found in {}".format(path))

    rows, report = [], []
    for f in files:
        wb = load_workbook(f, read_only=True, data_only=True)
        # The data sheet is not reliably called 'Combined' - one file names it
        # 'January 2026'. Take the first sheet that is not the disclaimer.
        sheet = next((s for s in wb.sheetnames
                      if "disclaimer" not in s.lower()), wb.sheetnames[0])
        data = list(wb[sheet].iter_rows(values_only=True))
        if not data:
            continue
        header = [str(h).strip() if h is not None else "" for h in data[0]]
        # Feb and March carry two stray trailing columns holding a count
        # parked in the header row. Keep only the columns we know.
        keep = {i: h for i, h in enumerate(header) if h in WANTED}
        n = 0
        for r in data[1:]:
            rec = {h: (str(r[i]).strip() if i < len(r) and r[i] is not None
                       else "") for i, h in keep.items()}
            if not rec.get("TIMBER_MARK"):
                continue
            rec["TIMBER_MARK"] = rec["TIMBER_MARK"].upper()
            rec["_file"] = os.path.basename(f)
            rec["_sheet"] = sheet
            rows.append(rec)
            n += 1
        report.append({
            "file": os.path.basename(f), "sheet": sheet, "rows": n,
            "marks": len({x["TIMBER_MARK"] for x in rows[-n:]}) if n else 0,
            "processed_on": next((x.get("ProcessedOn") for x in rows[-n:]
                                  if x.get("ProcessedOn")), ""),
            "stray_columns": len(header) - len(keep),
        })
    return rows, report


def union(rows: list[dict]) -> list[dict]:
    seen, out = set(), []
    for r in rows:
        key = (r["TIMBER_MARK"], r.get("PID", ""))
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


def marks_from_register(path: str, klass: str | None) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("--register needs openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in ("Register", "Sources_Detail")
                  if s in wb.sheetnames), wb.sheetnames[0])
    data = list(wb[sheet].iter_rows(values_only=True))
    header = [str(h or "").strip() for h in data[0]]
    low = [h.lower() for h in header]

    def find(*names):
        for n in names:
            if n in low:
                return low.index(n)
        return None

    i_id = find("corrected unit id", "unit id", "identifier")
    cands = [i for i, h in enumerate(low) if re.fullmatch(r"class\s*v?\d*", h)]
    i_cls = sorted(cands, reverse=True)[0] if cands else None
    if i_id is None:
        sys.exit("no identifier column on '{}'".format(sheet))
    out = []
    for r in data[1:]:
        if klass and i_cls is not None and i_cls < len(r):
            if str(r[i_cls] or "").strip().upper() != klass.upper():
                continue
        v = str(r[i_id] or "").strip().upper()
        if v and v not in out:
            out.append(v)
    return out


# ───────────────────────────── ParcelMap BC ────────────────────────────────

_pid_field: str | None = None


def pid_field() -> str:
    """Which field on layer 218 holds the PID. Discovered, not assumed."""
    global _pid_field
    if _pid_field is not None:
        return _pid_field
    try:
        fields = post(PARCELS, {"f": "json"}).get("fields", [])
    except Exception:
        _pid_field = ""
        return ""
    names = [f["name"] for f in fields]
    for want in ("PID", "PID_NUMBER", "PID_FORMATTED", "PARCEL_PID"):
        if want in names:
            _pid_field = want
            break
    else:
        _pid_field = next((n for n in names if "PID" in n.upper()), "")
    return _pid_field


def fetch_parcels(pids: list[str], log=None) -> tuple[list[dict], list[str]]:
    """Parcel polygons for a list of PIDs. Returns (features, not_found).

    Queried in batches with an IN clause. Both the zero-padded string and the
    integer form are tried, because the field type is not knowable in advance
    and getting it wrong looks exactly like a PID that does not exist.
    """
    field = pid_field()
    if not field:
        return [], list(pids)

    found: dict[str, dict] = {}
    todo = list(dict.fromkeys(pids))

    for form in ("string", "integer"):
        missing = [p for p in todo if p not in found]
        if not missing:
            break
        for i in range(0, len(missing), 100):
            batch = missing[i:i + 100]
            if form == "string":
                vals = ",".join("'{}'".format(p) for p in batch)
            else:
                vals = ",".join(str(int(p)) for p in batch)
            try:
                data = post(PARCELS + "/query", {
                    "where": "{} IN ({})".format(field, vals),
                    "outFields": "*", "returnGeometry": "true",
                    "outSR": 4326, "resultRecordCount": PAGE, "f": "geojson"})
            except Exception as exc:
                if log:
                    log("    {} batch failed: {}".format(form, str(exc)[:70]))
                continue
            for f in data.get("features", []):
                v = str((f.get("properties") or {}).get(field, "")).strip()
                key = re.sub(r"[^0-9]", "", v).zfill(9)
                if key not in found:
                    found[key] = f
            if log:
                log("    {} batch {}: {} of {} matched".format(
                    form, i // 100 + 1, len(found), len(todo)))
            time.sleep(0.2)

    return list(found.values()), [p for p in todo if p not in found]


def area_ha(feature: dict) -> float:
    p = feature.get("properties") or {}
    for k in ("FEATURE_AREA_SQM", "AREA_SQM", "SHAPE.AREA", "PARCEL_AREA"):
        v = p.get(k)
        if isinstance(v, (int, float)) and v > 0:
            return round(v / 10000.0, 2)
    return 0.0


# ──────────────────────────────── main ─────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("extracts", help="folder of the ministry .xlsx extracts")
    ap.add_argument("--marks", help="comma separated marks to resolve")
    ap.add_argument("--register", help="a HARP register to take marks from")
    ap.add_argument("--class", dest="klass", default="B",
                    help="class filter for --register (default B)")
    ap.add_argument("--inventory-only", action="store_true",
                    help="describe the extracts and stop")
    ap.add_argument("--out", default="ptm_output")
    ap.add_argument("--limit", type=int, help="cap the marks resolved")
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    print(ATTRIBUTION)

    rows, report = load_extracts(args.extracts)
    u = union(rows)
    print("\n" + "=" * 78)
    print("EXTRACTS")
    print("=" * 78)
    print("{:<46} {:>7} {:>7} {:>7}".format("FILE", "ROWS", "MARKS", "STRAY"))
    for r in report:
        print("{:<46} {:>7} {:>7} {:>7}".format(
            r["file"][:46], r["rows"], r["marks"], r["stray_columns"]))
    print("-" * 70)
    print("{:<46} {:>7} {:>7}".format("union, deduplicated on (mark, PID)",
                                      len(u), len({r["TIMBER_MARK"] for r in u})))
    print("\ndistricts: " + ", ".join(
        "{} {}".format(k, v) for k, v in
        Counter(r.get("ORG_UNIT_CODE", "") for r in u).most_common(8)))

    if args.inventory_only:
        return

    # ---- which marks ----
    if args.marks:
        wanted = [m.strip().upper() for m in args.marks.split(",") if m.strip()]
    elif args.register:
        wanted = marks_from_register(args.register, args.klass)
        print("\n{} marks from the register (class {})".format(
            len(wanted), args.klass))
    else:
        wanted = sorted({r["TIMBER_MARK"] for r in u})
        print("\nno --marks or --register given, resolving all {} marks in the "
              "extracts".format(len(wanted)))

    present = {r["TIMBER_MARK"] for r in u}
    matched = [m for m in wanted if m in present]
    missing = [m for m in wanted if m not in present]
    print("  {} of {} found in the extracts".format(len(matched), len(wanted)))
    if missing:
        print("  not present: " + ", ".join(missing[:12])
              + (" …" if len(missing) > 12 else ""))
    if args.limit:
        matched = matched[:args.limit]
    if not matched:
        print("\nNothing to resolve.")
        return

    # ---- PIDs ----
    print("\n" + "=" * 78)
    print("PARCELS")
    print("=" * 78)
    field = pid_field()
    print("  ParcelMap PID field: {}".format(field or "NOT FOUND - cannot query"))
    if not field:
        return

    by_mark: dict[str, dict] = {}
    bad: list[dict] = []
    for r in u:
        m = r["TIMBER_MARK"]
        if m not in matched:
            continue
        pids, note = parse_pids(r.get("PID"))
        if note and not pids:
            bad.append({"timber_mark": m, "raw_pid": r.get("PID"),
                        "note": note, "legal": r.get("LEGAL", "")[:120]})
        e = by_mark.setdefault(m, {"pids": [], "districts": set(),
                                   "legals": []})
        e["pids"].extend(pids)
        if r.get("ORG_UNIT_CODE"):
            e["districts"].add(r["ORG_UNIT_CODE"])
        if r.get("LEGAL"):
            e["legals"].append(r["LEGAL"])

    all_pids = sorted({p for e in by_mark.values() for p in e["pids"]})
    print("  {} marks  ·  {} distinct PIDs  ·  {} unparsable cells".format(
        len(by_mark), len(all_pids), len(bad)))

    feats, not_found = fetch_parcels(all_pids, log=print)
    by_pid = {re.sub(r"[^0-9]", "", str((f.get("properties") or {}).get(field, ""))
                     ).zfill(9): f for f in feats}
    print("  {} of {} PIDs returned a parcel".format(len(by_pid), len(all_pids)))

    # ---- per mark ----
    print("\n{:<10} {:>7} {:>7} {:>11} {:<10} {}".format(
        "MARK", "PIDS", "FOUND", "AREA HA", "DISTRICT", "LEGAL (first)"))
    print("-" * 100)
    out_features, summary = [], []
    for m in sorted(by_mark):
        e = by_mark[m]
        pids = sorted(set(e["pids"]))
        got = [by_pid[p] for p in pids if p in by_pid]
        total = round(sum(area_ha(f) for f in got), 1)
        for f in got:
            props = dict(f.get("properties") or {})
            props.update({
                "harp_timber_mark": m,
                "harp_pid": str(props.get(field, "")),
                "harp_source": "PMBC parcel via ministry scaled-timbermark extract",
                "harp_geometry_means": ("titled parcel the timber was scaled "
                                        "from - not the cut block boundary"),
                "harp_districts": ",".join(sorted(e["districts"])),
                "harp_retrieved": datetime.now().isoformat(timespec="seconds"),
                "harp_licence": ATTRIBUTION,
            })
            out_features.append({"type": "Feature",
                                 "geometry": f.get("geometry"),
                                 "properties": props})
        print("{:<10} {:>7} {:>7} {:>11,.1f} {:<10} {}".format(
            m, len(pids), len(got), total,
            ",".join(sorted(e["districts"]))[:10],
            (e["legals"][0] if e["legals"] else "")[:44]))
        summary.append({"timber_mark": m, "pids": len(pids),
                        "parcels_found": len(got),
                        "parcels_missing": len(pids) - len(got),
                        "area_ha": total,
                        "districts": ",".join(sorted(e["districts"])),
                        "example_legal": (e["legals"][0] if e["legals"] else "")})

    # ---- outputs ----
    gj = {"type": "FeatureCollection", "name": "ptm_parcels",
          "metadata": {
              "generated": datetime.now().isoformat(timespec="seconds"),
              "marks": len(by_mark), "pids": len(all_pids),
              "parcels": len(out_features), "licence": ATTRIBUTION,
              "note": ("A parcel is the titled land the timber was scaled "
                       "from. The harvest is somewhere inside it. Do not "
                       "present this as a harvest boundary.")},
          "features": out_features}
    gpath = os.path.join(args.out, "ptm_parcels_{}.geojson".format(stamp))
    with open(gpath, "w", encoding="utf-8") as fh:
        json.dump(gj, fh)

    spath = os.path.join(args.out, "ptm_summary_{}.csv".format(stamp))
    with open(spath, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(summary[0].keys()))
        w.writeheader(); w.writerows(summary)

    upath = os.path.join(args.out, "ptm_union_{}.csv".format(stamp))
    with open(upath, "w", encoding="utf-8-sig", newline="") as fh:
        cols = list(WANTED) + ["_file", "_sheet"]
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader(); w.writerows(u)

    if bad or not_found:
        qpath = os.path.join(args.out, "ptm_questions_{}.csv".format(stamp))
        with open(qpath, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["kind", "timber_mark", "value", "note"])
            for b in bad:
                w.writerow(["unparsable PID", b["timber_mark"], b["raw_pid"],
                            b["note"]])
            for p in not_found:
                w.writerow(["PID not in ParcelMap", "", p,
                            "parsed correctly but no parcel returned"])

    total_ha = round(sum(s["area_ha"] for s in summary), 1)
    print("\n" + "-" * 78)
    print("  {} marks  ·  {} parcels  ·  {:,.0f} ha of titled land".format(
        len(summary), len(out_features), total_ha))
    print("  {} PIDs found no parcel, {} cells could not be parsed".format(
        len(not_found), len(bad)))
    print("\n  A parcel is the land the timber was scaled from, not the cut.")
    print("  Treat the area above as an upper bound on the harvest.")
    print("\n  geometry : {}".format(gpath))
    print("  summary  : {}".format(spath))
    print("  union    : {}".format(upath))
    if bad or not_found:
        print("  questions: {}".format(qpath))
    print("-" * 78)


if __name__ == "__main__":
    main()

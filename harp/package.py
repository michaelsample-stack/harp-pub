"""Sorting a monthly package.

A client sends a folder. It holds a job list, one or more registry extracts,
and in time other things nobody has described yet. HARP has to work out which
is which without being told.

RECOGNISE BY COLUMNS, NOT BY FILENAME
-------------------------------------
Filenames in this data have already been proven wrong three separate ways: a
workbook named "June 2026" whose data sheet is "January 2026" and whose
records were processed in February; a "Calendar Year" label on files that are
not year-to-date; and a ProcessedOn field that varies per record rather than
per file. Nothing downstream should infer period, scope or completeness from
a filename.

Columns are reliable. SOURCEID and PRODUCT_TYPE is a job list. TIMBER_MARK and
PID is a private mark extract. That holds whatever the file is called.

TWO BEHAVIOURS, AND THE DIFFERENCE MATTERS
------------------------------------------
A job list REPLACES. It is the current statement of what needs answering, and
last month's is superseded.

A registry extract ACCUMULATES. The six private mark extracts are not
cumulative - the most recently processed one covers 27.7% of the year - so
taking the newest discards three quarters of the data. They are appended to a
store and deduplicated on (mark, PID), never replaced.

NOTHING IS DISCARDED SILENTLY
-----------------------------
A file that matches nothing is reported with the columns it actually had, so
the next kind of file to arrive announces itself rather than vanishing.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field

# A signature is a set of columns that must all be present, plus a set that
# must be absent. First match wins, so the more specific ones come first.
#
# The absent set is what stops a delivery record being read as a supply list.
# Both carry SOURCEID, and on columns alone the delivery summary matched - which
# would have put 2,543 delivery rows down the resolver as if each were a source.
SIGNATURES = [
    ("delivery_record", {"SOURCEID", "LOADID"}, set(),
     "a delivery record, not a supply list - one row per load, with volumes "
     "and dates. Read for volume, never resolved."),
    ("job_list", {"SOURCEID"}, {"LOADID", "BDT", "WS_TICKET"},
     "the client's own supply record - what needs answering. Replaces last "
     "month's."),
    ("private_marks", {"TIMBER_MARK", "PID"}, set(),
     "BC scaled-timbermark extract - a registry. Accumulates; never replaced."),
    ("lot_list", {"Lot ID"}, set(),
     "the month's production lots - what was made, when, how much, and the "
     "measured species split. The walkback starts here."),
    # Mill locations before the register: both carry a supplier column, and
    # the register's signature is the looser of the two.
    ("mill_locations", {"supplier", "latitude", "district_code"}, set(),
     "mill locations, from `harp mills`. Reference data for the search "
     "areas."),
    ("supplier_register", {"Supplier"}, {"Lot ID", "SOURCEID", "latitude"},
     "our own supplier register - says which suppliers still need a search "
     "area. Ours, not the client's."),
    ("supplier_geodata", {"geometry"}, set(),
     "geometry supplied for one source"),
]

# A JSON document has no columns, so it is recognised by its top-level keys.
DOCUMENT_SIGNATURES = [
    ("producer_geodata", {"Originator", "Products"},
     "harvest areas a producer declared in their own file - geometry, source "
     "id, timber marks, volumes and production dates. Taken at their word."),
    # No signature for a Digital Material Passport. They were a prior
    # third-party effort and nothing in the pipeline reads them, so one turning
    # up in a drop should be noticed rather than quietly filed under a kind
    # nobody consumes.
    ("harvest_units", {"__geometrycollection__"},
     "harvest units downloaded from a DMP. Collections and multiparts have to "
     "be exploded, and the polygons sorted into cutblocks and regional areas."),
]

# Anything with one of these extensions is worth opening
READABLE = (".xlsx", ".xlsm", ".csv", ".geojson", ".json")


@dataclass
class Item:
    path: str
    kind: str = "unknown"
    columns: list[str] = field(default_factory=list)
    sheet: str = ""
    rows: int = 0
    note: str = ""

    @property
    def accumulates(self) -> bool:
        """Whether a new copy adds to a store rather than replacing it."""
        return self.kind in ("private_marks", "harvest_units")

    @property
    def resolvable(self) -> bool:
        """Whether this file is something the resolver should be pointed at."""
        return self.kind == "job_list"


def _columns(path: str) -> tuple[list[str], str, int]:
    """Column names, sheet name and a row count. Empty on anything unreadable."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet = next((s for s in wb.sheetnames
                          if "disclaimer" not in s.lower()), wb.sheetnames[0])
            ws = wb[sheet]
            rows = list(ws.iter_rows(max_row=2, values_only=True))
            cols = [str(c).strip() for c in (rows[0] if rows else []) if c]
            return cols, sheet, ws.max_row or 0
        if ext == ".csv":
            import csv as _csv
            with open(path, encoding="utf-8-sig", newline="") as fh:
                r = _csv.reader(fh)
                cols = [c.strip() for c in next(r, [])]
                n = sum(1 for _ in r)
            return cols, "", n
        if ext in (".geojson", ".json"):
            import json as _json
            with open(path, encoding="utf-8") as fh:
                blob = _json.load(fh)
            if not isinstance(blob, dict):
                return [], "", 0
            if blob.get("type") == "FeatureCollection":
                feats = blob.get("features") or []
                cols = ["geometry"]
                if feats:
                    cols += sorted((feats[0].get("properties") or {}).keys())[:20]
                    # A GeometryCollection is the tell for a DMP harvest units
                    # download. It is also not valid under EUDR, so noticing it
                    # here saves a validation failure later.
                    if any((f.get("geometry") or {}).get("type")
                           == "GeometryCollection" for f in feats[:5]):
                        cols.append("__geometrycollection__")
                return cols, "", len(feats)
            # any other JSON document - recognised by its top-level keys
            return sorted(blob.keys()), "", 0
    except Exception:
        pass
    return [], "", 0


def classify(path: str) -> Item:
    cols, sheet, rows = _columns(path)
    item = Item(path=path, columns=cols, sheet=sheet, rows=rows)
    if not cols:
        item.note = "could not read this file"
        return item
    upper = {c.upper() for c in cols}

    for kind, keys, note in DOCUMENT_SIGNATURES:
        if {k.upper() for k in keys} <= upper:
            item.kind = kind
            item.note = note
            return item

    for kind, required, forbidden, note in SIGNATURES:
        if not {r.upper() for r in required} <= upper:
            continue
        if {f.upper() for f in forbidden} & upper:
            continue
        item.kind = kind
        item.note = note
        return item

    item.note = "no signature matched - columns: " + ", ".join(cols[:12])
    return item


def sort_package(folder: str) -> dict[str, list[Item]]:
    """Every readable file in a folder, grouped by what it turned out to be."""
    out: dict[str, list[Item]] = {}
    if not os.path.isdir(folder):
        return out
    for name in sorted(os.listdir(folder)):
        path = os.path.join(folder, name)
        if not os.path.isfile(path):
            continue
        if not name.lower().endswith(READABLE):
            continue
        if name.startswith("~$"):        # an Excel lock file
            continue
        item = classify(path)
        out.setdefault(item.kind, []).append(item)
    return out


def describe(sorted_items: dict[str, list[Item]]) -> str:
    if not sorted_items:
        return "no readable files found"
    lines = []
    for kind in sorted(sorted_items):
        items = sorted_items[kind]
        lines.append("{}  ({} file{})".format(
            kind, len(items), "" if len(items) == 1 else "s"))
        for it in items:
            lines.append("    {:<52} {:>7} rows{}".format(
                os.path.basename(it.path)[:52], it.rows,
                "  [{}]".format(it.sheet) if it.sheet else ""))
        if kind in ("unknown", "delivery_record", "dmp", "harvest_units"):
            for it in items:
                lines.append("      {}".format(it.note[:110]))
    return "\n".join(lines)

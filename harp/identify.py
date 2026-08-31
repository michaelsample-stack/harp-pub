"""Reading a raw supply list, and deciding what each identifier might be.

Local. No network. Two jobs:

    load()      turn whatever the client sent into Record objects
    shapes()    rank what an identifier could be
    klass()     how many parties stand between it and a harvest area

On ranking rather than routing
------------------------------
Codes like '61/243' look like a forest file id and a cutting permit, and are in
fact timber marks verbatim, slash included. Committing to a shape and querying
only the matching field skipped the field that held them and produced seven
false negatives on the first Harmac run. So shapes() proposes an ordered list
and the router tries them in order. Nothing is eliminated on appearance.

On the client's list changing
-----------------------------
A supplier register drifts constantly. Nothing here depends on a curated one.
The minimum viable input is an identifier; a jurisdiction and a product type
make the answer better but are not required.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any, Iterable

from .resolution import IdShape, Klass

# Values that appear in an identifier column but are not identifiers
NOT_IDENTIFIERS = {
    "TRUCKED", "WATERED", "TIMBERSALE", "TRADE", "OTHER", "HOG", "CHIPPER",
    "EMILL", "LANDFILL", "F", "N/A", "NONE",
}

# Words that mark an aggregation point rather than a mill
YARD_WORDS = ("YARD", "RELOAD", "TRADE", "SORT", "BARGE", "SCOW", "DUMP",
              "RL", "CHIP-RELOAD")

CHIP_PRODUCTS = {"BULK", "CHIP", "CHIPS", "HOG", "RESIDUAL"}

STATE_TO_JURISDICTION = {
    "BC": "BC", "BRITISH COLUMBIA": "BC",
    "WA": "WA", "WASHINGTON": "WA",
    "OR": "OR", "OREGON": "OR",
    "AK": "AK", "ALASKA": "AK",
    "CA": "CA", "CALIFORNIA": "CA",
    "CA-BC": "BC", "US-WA": "WA", "US-OR": "OR", "US-AK": "AK", "US-CA": "CA",
}

# Column names seen in the wild, in preference order
ID_COLS = ("corrected unit id", "unit id", "unitid", "identifier",
           "timber mark", "timber_mark", "mark", "timbermark")
SRC_COLS = ("source id", "sourceid", "source_id")
# Deliberately not a bare "name". SOURCE.xlsx has a NAME column holding a
# delivery description - "Direct Delivery DCT", "SB4 LADYSMITH-DCT" - which is
# not the supplier. Matching it put a haul arrangement in the supplier field
# on every resolved feature.
SUP_COLS = ("supplier name", "supp_name", "supplier", "client_name")
SUPID_COLS = ("supplier id", "suppid", "supplier_id")
PROD_COLS = ("product type", "product_type", "productid", "product")
JUR_COLS = ("jurisdiction", "province / state", "province", "state",
            "stateid", "state code")


@dataclass
class Record:
    """One row of a client's supply list."""

    source_id: str
    identifier: str
    supplier_name: str = ""
    supplier_id: str = ""
    jurisdiction: str = ""            # 'BC' | 'WA' | 'OR' | 'AK' | 'CA'
    product_type: str = ""
    klass: Klass | None = None
    raw: dict[str, Any] = field(default_factory=dict)


# ─────────────────────────────── identify ──────────────────────────────────

def shapes(identifier: str) -> list[IdShape]:
    """Candidate readings, most likely first. Never empty."""
    u = str(identifier or "").strip().upper()
    if not u:
        return [IdShape.UNKNOWN]

    out: list[IdShape] = []

    def add(s: IdShape) -> None:
        if s not in out:
            out.append(s)

    # A timber mark goes first whatever the code looks like. In practice it is
    # the only field that has ever matched - 71 of 71 resolutions on the
    # Harmac data - so trying anything else first only wastes a query.
    if u not in NOT_IDENTIFIERS and (re.fullmatch(r"[A-Z0-9'/-]{4,10}", u)):
        add(IdShape.TIMBER_MARK)
    if re.fullmatch(r"A\d{4}", u):
        add(IdShape.LICENCE)
    if "/" in u:
        add(IdShape.CUTTING_PERMIT)
    if u.startswith("DSI"):
        add(IdShape.DRYLAND_SORT)
    if u in NOT_IDENTIFIERS or any(w in u for w in YARD_WORDS):
        add(IdShape.COMPANY_NAME)
    if " " in u:
        add(IdShape.PLACE_NAME)
    if not out:
        add(IdShape.UNKNOWN)
    return out


def klass(record: Record) -> Klass:
    """Assign a class if the caller has not.

    THE IDENTIFIER DECIDES, NOT THE PRODUCT TYPE
    --------------------------------------------
    An earlier version classified every chip source as C2 on the grounds that
    chip sources carry mill names. Most do - but not all. Three identifiers in
    the Harmac data sit on both a log source and a chip source, and one chip
    source carries a licence number that resolves to a cut block. Classifying
    on product type skipped those, and silently lost a genuine plot-level
    match: the causation was backwards.

    Product type is a hint about what the identifier is likely to be, not a
    fact about how many parties stand in the way. So: if the value looks like
    a harvest identifier it is tried, whatever product it arrived against. The
    cost of being wrong is one query; the cost of not trying is a lost cut
    block.

    Product type only breaks the tie for values that are not harvest
    identifiers - deciding whether a mill name is a processor (C2) or an
    aggregation point (D).

    This stage answers only the first question - does the record carry a
    harvest identifier. The router answers the second, and may revise A to B
    once it knows whether a public register holds the geometry.
    """
    if record.klass:
        return record.klass

    u = str(record.identifier or "").strip().upper()
    product = (record.product_type or "").upper()

    if not u:
        return Klass.E

    # A licence number is unambiguous - A2711 is not a place. Anything shaped
    # like one is class A whatever product it arrived against, which is what
    # rescued A2500 and 99095 from being written off as mill names.
    if u not in NOT_IDENTIFIERS and IdShape.LICENCE in shapes(u):
        return Klass.A

    # Otherwise product type is the better signal for the CLASS. A chip source
    # usually names a mill or a yard. But note this decides only what we
    # believe about the supply tier - it does not decide whether to query.
    # See `worth_querying` below and the router's short-circuit.
    if product in CHIP_PRODUCTS:
        return (Klass.D if (any(w in u for w in YARD_WORDS)
                            or u in NOT_IDENTIFIERS) else Klass.C2)

    if u in NOT_IDENTIFIERS:
        return Klass.E
    sh = shapes(u)
    if IdShape.TIMBER_MARK in sh or IdShape.LICENCE in sh:
        return Klass.A          # provisional - the router may revise it
    return Klass.E


def worth_querying(identifier: str) -> bool:
    """Whether this value looks like it could be a harvest identifier.

    A HINT ONLY. Nothing is filtered on it - every source runs the full
    ladder regardless. It exists so a caller can report what it expects, and
    so a surprising result can be recognised as surprising.

    The reason it does not gate anything: a mill name and a timber mark are
    not distinguishable by shape - TACOMA and EDRWD are both six alphabetic
    characters - and the codes that mattered most in practice were all ones
    judged unpromising by eye and found to be real when finally tested.
    """
    u = str(identifier or "").strip().upper()
    if not u or u in NOT_IDENTIFIERS:
        return False
    sh = shapes(u)
    return IdShape.TIMBER_MARK in sh or IdShape.LICENCE in sh


def identify(record: Record) -> Record:
    record.klass = klass(record)
    return record


# ──────────────────────────────── loading ──────────────────────────────────

def _pick_class_column(lower: list[str]) -> int | None:
    """A register carries several columns starting with 'class' - 'Class Basis'
    among them - so an exact versioned name wins and prose lookalikes are
    excluded outright."""
    versioned = sorted((h for h in lower if re.fullmatch(r"class\s*v?\d+", h)),
                       key=lambda h: int(re.search(r"\d+", h).group()),
                       reverse=True)
    if versioned:
        return lower.index(versioned[0])
    if "class" in lower:
        return lower.index("class")
    for h in lower:
        if h.startswith("class") and not any(
                w in h for w in ("basis", "note", "reason", "family",
                                 "model", "description", "summary")):
            return lower.index(h)
    return None


def _map_columns(header: list[str]) -> dict[str, str | None]:
    lower = [str(h or "").strip().lower() for h in header]

    def find(names: Iterable[str]) -> str | None:
        for n in names:
            if n in lower:
                return header[lower.index(n)]
        return None

    ic = _pick_class_column(lower)
    return {
        "id": find(ID_COLS) or (header[0] if header else None),
        "source": find(SRC_COLS),
        "supplier": find(SUP_COLS),
        "supplier_id": find(SUPID_COLS),
        "product": find(PROD_COLS),
        "jurisdiction": find(JUR_COLS),
        "class": header[ic] if ic is not None else None,
    }


def _is_label_row(row: dict, cols: dict) -> bool:
    """Whether this row is a second header rather than data.

    A LIMS export often carries machine names on row 1 and human labels on
    row 2 - SOURCE.xlsx does exactly that. Read naively, the label row becomes
    a record with identifier 'Unit ID', which then goes off to a government
    register and comes back unresolved. It is caught by the values echoing the
    column names.
    """
    hits = 0
    for key in ("id", "source", "supplier", "product"):
        col = cols.get(key)
        if not col:
            continue
        val = str(row.get(col, "") or "").strip().lower()
        if not val:
            continue
        if val == str(col).strip().lower() or val.replace(" ", "_") == \
                str(col).strip().lower().replace(" ", "_"):
            hits += 1
    return hits >= 2


def _record(row: dict, cols: dict, default_jurisdiction: str) -> Record:
    def g(key: str) -> str:
        c = cols.get(key)
        return str(row.get(c, "") or "").strip() if c else ""

    ident = g("id")
    k = None
    if g("class"):
        try:
            k = Klass(g("class").upper() if g("class").upper() != "N/A" else "N/A")
        except ValueError:
            k = None
    jur = STATE_TO_JURISDICTION.get(g("jurisdiction").upper(),
                                    default_jurisdiction)
    return Record(
        source_id=g("source") or ident,
        identifier=ident,
        supplier_name=g("supplier"),
        supplier_id=g("supplier_id"),
        jurisdiction=jur,
        product_type=g("product"),
        klass=k,
        raw={k2: v for k2, v in row.items() if v not in (None, "")},
    )


LAST_MAPPING: dict[str, str | None] = {}


def load(path: str, sheet: str | None = None, only_class: str | None = None,
         default_jurisdiction: str = "BC") -> list[Record]:
    """Read a raw supply list. CSV or XLSX, whatever columns it happens to have.

    The column mapping it chose is left in LAST_MAPPING so a caller can report
    it. Silent mis-mapping is the failure mode that matters here: if the client
    renames a column between monthly drops, the loader will pick something else
    and every downstream number will be quietly wrong.
    """
    if str(path).lower().endswith((".xlsx", ".xlsm")):
        return _load_xlsx(path, sheet, only_class, default_jurisdiction)
    return _load_csv(path, only_class, default_jurisdiction)


def describe_mapping() -> str:
    if not LAST_MAPPING:
        return "no file loaded"
    return "  ".join("{}={}".format(k, v or "-")
                     for k, v in LAST_MAPPING.items() if k != "_sheet")


def _load_csv(path, only_class, default_jurisdiction) -> list[Record]:
    from . import io
    rows = io.read_csv_dicts(path)
    if not rows:
        return []
    cols = _map_columns(list(rows[0].keys()))
    LAST_MAPPING.clear(); LAST_MAPPING.update(cols)
    out = []
    for r in rows:
        if _is_label_row(r, cols):
            continue
        if only_class and cols["class"]:
            if str(r.get(cols["class"], "")).strip().upper() != only_class.upper():
                continue
        rec = _record(r, cols, default_jurisdiction)
        if rec.identifier:
            out.append(rec)
    return out


def _load_xlsx(path, sheet, only_class, default_jurisdiction) -> list[Record]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xlsx needs openpyxl:  pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    name = sheet or ("Sources_Detail" if "Sources_Detail" in wb.sheetnames
                     else wb.sheetnames[0])
    rows = list(wb[name].iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    cols = _map_columns(header)
    LAST_MAPPING.clear(); LAST_MAPPING.update(cols)
    LAST_MAPPING["_sheet"] = name

    out, seen_classes = [], set()
    for r in rows[1:]:
        row = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if _is_label_row(row, cols):
            continue
        if cols["class"] and row.get(cols["class"]):
            seen_classes.add(str(row[cols["class"]]).strip())
        if only_class and cols["class"]:
            if str(row.get(cols["class"], "")).strip().upper() != only_class.upper():
                continue
        rec = _record(row, cols, default_jurisdiction)
        if rec.identifier:
            out.append(rec)

    if only_class and not out and seen_classes:
        raise RuntimeError(
            "No rows with class '{}' on sheet '{}'. Column read: '{}'. "
            "Values present: {}".format(only_class, name, cols["class"],
                                        ", ".join(sorted(seen_classes))))
    return out


# Preference order when the same identifier arrives on several sources: try
# the one most likely to resolve. Ties are broken by order of appearance.
_KEEP_ORDER = {Klass.A: 0, Klass.B: 1, Klass.E: 2, Klass.C1: 3,
               Klass.C2: 4, Klass.D: 5, Klass.NA: 6}


def dedupe(records: list[Record]) -> list[Record]:
    """One resolution per identifier per jurisdiction.

    Twelve of Harmac's identifiers serve more than one source. Resolving each
    once and fanning the answer back out saves both queries and double-counted
    block totals.

    Which of the duplicates is kept matters. Three identifiers sit on both a
    log source and a chip source, and keeping whichever happened to come first
    once caused a licence number to be skipped as a mill name. The record with
    the most promising class wins.
    """
    best: dict[tuple, Record] = {}
    order: list[tuple] = []
    for r in records:
        key = (r.identifier.upper(), r.jurisdiction)
        rank = _KEEP_ORDER.get(klass(r), 9)
        if key not in best:
            best[key] = r
            order.append(key)
        elif rank < _KEEP_ORDER.get(klass(best[key]), 9):
            best[key] = r
    return [best[k] for k in order]

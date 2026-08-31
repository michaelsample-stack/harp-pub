"""Where each supplier operates - mill point and administrative district.

One question, two uses. A supplier's operating area is the district filter for
a tenure-register lookup, and it is the catchment when there is no register
match - so it is one piece of work rather than two.

    harp mills <supplier register.xlsx>

WHERE IT LOOKS
--------------
    BC Major Timber Processing Facilities, layer 793
        252 mills as points, with company name and annual capacity. Free, and
        it covers most of the British Columbia supply base.

    Natural Resource Districts, layer 748
        the point is intersected with this to get the district a mill sits in,
        which is the filter a register lookup needs.

    The supplier's own name
        many carry a place: "Campbell River Shake", "Comox Valley Shakes",
        "Cinnabar Valley". Matched against a district name where the facility
        list has nothing.

WHAT IT CANNOT DO
-----------------
A broker with no mill will not appear in a facility list, because there is no
facility. Those come back empty and need looking up by hand - which is a
finding, not a failure, and is why the output carries how each answer was
arrived at.

ON PROVENANCE
-------------
Every row records how the location was established: a facility match, a place
name in the supplier's own name, or nothing. A district a supplier told us and
one inferred from a mill point are not equal evidence, and a column that
records which costs nothing now and is impossible to reconstruct later.

A MILL IS NOT A HARVEST AREA
----------------------------
The mill is where the chips were made. Its district narrows a search; it does
not say where the wood grew. A mill in Nanaimo may buy logs from four
districts.
"""

from __future__ import annotations

import csv
import json
import os
import re
import time
from datetime import datetime

import requests

ROOT = ("https://delivery.maps.gov.bc.ca/arcgis/rest/services"
        "/mpcm/bcgwpub/MapServer")
FACILITIES = 793
DISTRICTS = 748

TIMEOUT = 120
S = requests.Session()
S.headers.update({"User-Agent": "NGIS-HARP-mills/1.0"})

# Words that carry no company identity. Matching on these alone finds the
# wrong firm - "Cedar" matched three different suppliers to Teal Cedar
# Products in the envelope run.
GENERIC = {
    "LTD", "LTD.", "INC", "INC.", "LIMITED", "CORP", "CORP.", "CORPORATION",
    "COMPANY", "CO", "CO.", "LP", "LLP", "GP", "THE", "AND", "&", "GROUP",
    "HOLDINGS", "PARTNERSHIP", "ENTERPRISES", "INDUSTRIES", "PRODUCTS",
    "FOREST", "FORESTRY", "FORESTS", "LOGGING", "LUMBER", "TIMBER", "WOOD",
    "WOODS", "FIBRE", "FIBER", "MILL", "MILLS", "SAWMILL", "SAWMILLS",
    "SHAKE", "SHAKES", "SHINGLE", "CEDAR", "FIR", "HEMLOCK", "SPRUCE", "PINE",
    "CONTRACTING", "SERVICES", "MANAGEMENT", "RESOURCE", "RESOURCES",
    "TRADING", "SUPPLY", "CHIPPING", "CUSTOM", "SPLIT", "BARGE",
}


def post(url: str, params: dict) -> dict:
    r = S.post(url, data={**params, "f": "json"}, timeout=TIMEOUT)
    r.raise_for_status()
    data = r.json()
    if "error" in data:
        raise RuntimeError(data["error"].get("message", str(data["error"])))
    return data


def sql(v) -> str:
    return str(v).replace("'", "''")


def clean_name(name: str) -> str:
    """Strip the parenthetical qualifiers Harmac's records carry.

    'Alta Forest Products (Morton)' and 'Halo Sawmill (split barge)' are the
    same company as their unqualified forms - the bracket is a delivery
    arrangement, not a different firm.
    """
    n = re.sub(r"\(.*?\)", " ", str(name or "")).upper()
    n = re.sub(r"[^A-Z0-9&\s'-]", " ", n)
    return re.sub(r"\s+", " ", n).strip()


def distinctive(name: str) -> list[str]:
    """The words that actually identify a company."""
    return [w for w in clean_name(name).split()
            if w not in GENERIC and len(w) >= 3]


# ─────────────────────────── the facility list ─────────────────────────────

_facilities: list[dict] | None = None


def facilities() -> list[dict]:
    """Every BC timber processing facility, fetched once.

    252 records, so the whole list is pulled and matched locally rather than
    querying per supplier. Faster, and it lets a near-miss be reported instead
    of silently returning nothing.
    """
    global _facilities
    if _facilities is not None:
        return _facilities
    try:
        meta = post("{}/{}".format(ROOT, FACILITIES), {})
        names = [f["name"] for f in meta.get("fields", [])]
    except Exception:
        _facilities = []
        return _facilities

    name_field = next((n for n in names
                       if "COMPANY" in n.upper() or "NAME" in n.upper()), None)
    want = [n for n in names if n in (
        name_field, "LATITUDE", "LONGITUDE", "MILL_TYPE_DESC",
        "EST_AN_CAP_000_BDUS", "EST_AN_CAP_000_TONNES", "CITY", "OWNER_NAME",
        "FACILITY_NAME", "MILL_STATUS", "MILL_STATUS_DESC")]
    out = []
    try:
        data = post("{}/{}/query".format(ROOT, FACILITIES),
                    {"where": "1=1", "outFields": ",".join(want) or "*",
                     "returnGeometry": "true", "outSR": 4326,
                     "resultRecordCount": 1000})
    except Exception:
        _facilities = []
        return _facilities

    for f in data.get("features", []):
        a = f.get("attributes", {})
        g = f.get("geometry") or {}
        lat = a.get("LATITUDE") or g.get("y")
        lon = a.get("LONGITUDE") or g.get("x")
        label = " ".join(str(a.get(k) or "") for k in
                         (name_field, "OWNER_NAME", "FACILITY_NAME")).strip()
        out.append({"raw": a, "label": label, "lat": lat, "lon": lon,
                    "city": a.get("CITY", ""),
                    "type": a.get("MILL_TYPE_DESC", ""),
                    "capacity_bdu": a.get("EST_AN_CAP_000_BDUS")})
    _facilities = out
    return out


def match_facility(name: str) -> tuple[dict | None, str]:
    """A facility for this supplier, and how confident the match is.

    Full name first, then every distinctive word together, then a single
    distinctive word. A single GENERIC word is never enough - that is what
    matched three suppliers to the same company last time.
    """
    facs = facilities()
    if not facs:
        return None, ""
    target = clean_name(name)
    if not target:
        return None, ""

    for f in facs:
        if clean_name(f["label"]) == target:
            return f, "exact name"

    words = distinctive(name)
    if not words:
        return None, ""

    # every distinctive word present
    hits = [f for f in facs
            if all(w in clean_name(f["label"]) for w in words)]
    if len(hits) == 1:
        return hits[0], "all distinctive words"
    if len(hits) > 1:
        return hits[0], "all distinctive words ({} candidates)".format(len(hits))

    # the longest single distinctive word, only if it is reasonably specific
    for w in sorted(words, key=len, reverse=True):
        if len(w) < 5:
            continue
        hits = [f for f in facs if w in clean_name(f["label"])]
        if len(hits) == 1:
            return hits[0], "one distinctive word: {}".format(w)
        if len(hits) > 1:
            return None, "ambiguous on '{}' — {} facilities".format(w, len(hits))
    return None, ""


# ────────────────────────────── districts ──────────────────────────────────

_districts: list[dict] | None = None


def districts() -> list[dict]:
    global _districts
    if _districts is not None:
        return _districts
    out = []
    try:
        data = post("{}/{}/query".format(ROOT, DISTRICTS),
                    {"where": "1=1",
                     "outFields": "DISTRICT_NAME,ORG_UNIT,REGION_ORG_UNIT_NAME",
                     "returnGeometry": "false", "resultRecordCount": 200})
        for f in data.get("features", []):
            a = f["attributes"]
            out.append({"name": a.get("DISTRICT_NAME", ""),
                        "code": a.get("ORG_UNIT", ""),
                        "region": a.get("REGION_ORG_UNIT_NAME", "")})
    except Exception:
        pass
    _districts = out
    return out


def district_at(lat: float, lon: float) -> dict | None:
    """The natural resource district containing a point."""
    if lat is None or lon is None:
        return None
    try:
        data = post("{}/{}/query".format(ROOT, DISTRICTS), {
            "geometry": json.dumps({"x": lon, "y": lat,
                                    "spatialReference": {"wkid": 4326}}),
            "geometryType": "esriGeometryPoint",
            "spatialRel": "esriSpatialRelIntersects", "inSR": 4326,
            "outFields": "DISTRICT_NAME,ORG_UNIT,REGION_ORG_UNIT_NAME",
            "returnGeometry": "false", "resultRecordCount": 2})
    except Exception:
        return None
    feats = data.get("features") or []
    if not feats:
        return None
    a = feats[0]["attributes"]
    return {"name": a.get("DISTRICT_NAME", ""), "code": a.get("ORG_UNIT", ""),
            "region": a.get("REGION_ORG_UNIT_NAME", "")}


def district_from_name(name: str) -> dict | None:
    """A district implied by a place name in the supplier's own name.

    'Campbell River Shake', 'Comox Valley Shakes'. Weaker than a facility
    match and recorded as such, but better than nothing for a broker with no
    mill.
    """
    words = clean_name(name).split()
    if not words:
        return None
    for d in districts():
        dn = clean_name(d["name"]).replace("NATURAL RESOURCE DISTRICT", "").strip()
        if not dn:
            continue
        parts = [p for p in dn.split() if p not in GENERIC and len(p) >= 4]
        if parts and all(p in words for p in parts):
            return d
    return None


# ──────────────────────────────── inputs ───────────────────────────────────

def suppliers_from_register(path: str, klass: str | None) -> list[str]:
    """Supplier names from a register. Kept for callers that want names only."""
    return [name for name, _j in suppliers_with_jurisdiction(path, klass)]


def suppliers_with_jurisdiction(path: str,
                                klass: str | None) -> list[tuple]:
    """Supplier names, each with the jurisdiction the register gives it.

    The jurisdiction matters because the BC facility list will happily match
    a company name that also operates in BC. Weyerhaeuser and Interfor both
    do, and both were placed in BC districts on the strength of the name when
    the source Harmac buys from is in Washington.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("reading a register needs openpyxl:  "
                           "pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    preferred = ("Suppliers", "Catchments_needed", "All_suppliers", "Register")
    order = [s for s in preferred if s in wb.sheetnames]
    order += [s for s in wb.sheetnames if s not in order]

    for sheet in order:
        rows = list(wb[sheet].iter_rows(values_only=True))
        for header_row in range(0, min(6, len(rows))):
            low = [str(h or "").strip().lower() for h in rows[header_row]]
            i_sup = next((low.index(n) for n in
                          ("supplier", "supplier name", "name") if n in low),
                         None)
            if i_sup is None:
                continue
            i_cls = next((i for i, h in enumerate(low)
                          if h == "class" or re.fullmatch(r"class\s*v?\d*", h)),
                         None)
            i_jur = next((i for i, h in enumerate(low)
                          if h in ("jurisdiction", "province / state",
                                   "province/state", "state", "province",
                                   "region", "where")), None)
            out, seen = [], set()
            for r in rows[header_row + 1:]:
                if i_sup >= len(r) or not r[i_sup]:
                    continue
                if klass and i_cls is not None and i_cls < len(r):
                    cell = str(r[i_cls] or "").upper()
                    if klass.upper() not in [c.strip() for c in cell.split(",")]:
                        continue
                v = str(r[i_sup]).strip()
                jur = ""
                if i_jur is not None and i_jur < len(r):
                    jur = str(r[i_jur] or "").strip()
                if v and v not in seen:
                    seen.add(v)
                    out.append((v, jur))
            if out:
                return out
    raise RuntimeError("No supplier column found in {}".format(
        os.path.basename(path)))


# Anything not British Columbia. The BC facility list and the BC district
# layer are both BC-only, so placing a supplier from any of these in a BC
# district is wrong however well the company name matches.
NON_BC = {"WA", "WASHINGTON", "OR", "OREGON", "CA", "CALIFORNIA",
          "AK", "ALASKA", "ID", "IDAHO", "MT", "MONTANA", "US", "USA"}


def is_bc(jurisdiction: str) -> bool:
    j = (jurisdiction or "").strip().upper()
    if not j:
        return True          # unstated: try, and let the match speak
    if j in NON_BC:
        return False
    return not any(x in j.split() for x in NON_BC)


# ──────────────────────────────── main ─────────────────────────────────────

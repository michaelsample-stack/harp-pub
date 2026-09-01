"""Search areas — one for every source that no identifier could place.


WHAT A CATCHMENT IS HERE
------------------------
An area to search inside, not an answer. Every polygon in this layer is a
bounded region within which a supplier's fibre plausibly originated; change
detection then finds the ground that was actually disturbed. Nothing here is a
directly traced and nothing here should reach a declaration unrefined.

That is the pattern the other engagements use. Enviva buffers a mill by a
configurable radius and then queries harvest polygons inside it - the buffer is
a filter, never the declared area. Domtar joins supplier-declared county names
to published boundaries. Billerud sizes a buffer by the volume a source
produced. All three bound a search and declare what is found within it.

FOUR METHODS, BEST FIRST
------------------------
    1  operator tenure     the company's own cut blocks, from FTEN. Real
                           harvest geometry, just more of it than this client
                           bought.
    2  named district      the natural resource district the mill sits in, as
                           a published boundary. The Domtar pattern.
    3  mill buffer         a circle round the mill. The Enviva pattern, and
                           the weakest thing here.
    4  none                no mill, no tenure, no district. Recorded as a gap
                           rather than invented - Domtar returns null geometry
                           for a supplier who says "potentially all counties",
                           and an unbounded answer is better logged as no
                           answer.

RADIUS
------
Per supplier, not fixed. Enviva's config carries buffer_miles per mill because
a coastal barge haul and an interior truck haul are different economics. The
default here is 150 km, which is a starting position and not a measurement -
it should be replaced with something derived as soon as there is anything to
derive it from.

The buffer is built in a local azimuthal equidistant projection so it is a
true circle on the ground rather than a distorted one in degrees.
"""

from __future__ import annotations

import csv
import json
import os
import re
import time
from datetime import datetime

import requests

ROOT = ("https://delivery.maps.gov.bc.ca/arcgis/rest/services"
        "/mpcm/bcgwpub/MapServer")
BLOCKS, DISTRICTS = 340, 748
TIMEOUT = 180

# A source identifier is usually the mill town. It is not a harvest location -
# Mercer mills at Castlegar and draws from across the Kootenays - but it places
# the operation, and a named district is a better answer than a circle.
#
# This is Domtar's method with a weaker input. They used areas the supplier
# declared; this infers one from where the mill sits. The provenance field
# carries that difference so the two never look alike.
TOWN_DISTRICT = {
    "PARKSVILLE": ("DSI", "South Island"),
    "DUNCAN": ("DSI", "South Island"),
    "NANAIMO": ("DSI", "South Island"),
    "CHEMAINUS": ("DSI", "South Island"),
    "LADYSMITH": ("DSI", "South Island"),
    "PORT ALBERNI": ("DSI", "South Island"),
    "PORTALBERNI": ("DSI", "South Island"),
    "ERRINGTON": ("DSI", "South Island"),
    "DUKE": ("DSI", "South Island"),
    "COURTENAY": ("DCR", "Campbell River"),
    "CAMPBELL RIVER": ("DCR", "Campbell River"),
    "GOLD RIVER": ("DCR", "Campbell River"),
    "BLACKCREEK": ("DCR", "Campbell River"),
    "BLACK CREEK": ("DCR", "Campbell River"),
    "PORT MCNEILL": ("DNI", "North Island - Central Coast"),
    "PORT HARDY": ("DNI", "North Island - Central Coast"),
    "ABBOTSFORD": ("DCK", "Chilliwack"),
    "MISSION": ("DCK", "Chilliwack"),
    "MAPLE RIDGE": ("DCK", "Chilliwack"),
    "MAPLE-RIDGE": ("DCK", "Chilliwack"),
    "PITT MEADOWS": ("DCK", "Chilliwack"),
    "SURREY": ("DCK", "Chilliwack"),
    "DELTA": ("DCK", "Chilliwack"),
    "RICHMOND": ("DCK", "Chilliwack"),
    "SILVERDALE": ("DCK", "Chilliwack"),
    "LANGLEY": ("DCK", "Chilliwack"),
    "MERRITT": ("DCC", "Cascades"),
    "PRINCETON": ("DCC", "Cascades"),
    "KAMLOOPS": ("DKA", "Thompson Rivers"),
    "CASTLEGAR": ("DSE", "Selkirk"),
    "REVELSTOKE": ("DSE", "Selkirk"),
    "KELOWNA": ("DOS", "Okanagan Shuswap"),
    "VERNON": ("DOS", "Okanagan Shuswap"),
    "TERRACE": ("DKM", "Coast Mountains"),
    "SMITHERS": ("DND", "Nadina"),
    "QUESNEL": ("DQU", "Quesnel"),
    "PRINCE GEORGE": ("DPG", "Prince George"),
}

# Starting radius. Not a measurement - see the module docstring.
DEFAULT_RADIUS_KM = 150.0
KM = 1000.0

S = requests.Session()
S.headers.update({"User-Agent": "NGIS-HARP-catchments/1.0"})

ATTRIBUTION = ("Contains information licensed under the Open Government "
               "Licence - British Columbia.")


def post(url: str, params: dict) -> dict:
    r = S.post(url, data={**params, "f": "json"}, timeout=TIMEOUT)
    r.raise_for_status()
    d = r.json()
    if "error" in d:
        raise RuntimeError(d["error"].get("message", str(d["error"])))
    return d


def sql(v) -> str:
    return str(v).replace("'", "''")


# ─────────────────────────── method 3: buffer ──────────────────────────────

def mill_buffer(lat: float, lon: float, radius_km: float) -> dict | None:
    """A true circle on the ground, in WGS84.

    Projected to a local azimuthal equidistant frame first. Buffering degrees
    directly gives an ellipse that is wrong by a third at this latitude.
    """
    try:
        import pyproj
        from shapely.geometry import Point, mapping
        from shapely.ops import transform
    except ImportError:
        return None
    aeqd = pyproj.CRS.from_proj4(
        "+proj=aeqd +lat_0={} +lon_0={} +x_0=0 +y_0=0 +datum=WGS84 "
        "+units=m +no_defs".format(lat, lon))
    wgs = pyproj.CRS.from_epsg(4326)
    fwd = pyproj.Transformer.from_crs(wgs, aeqd, always_xy=True).transform
    back = pyproj.Transformer.from_crs(aeqd, wgs, always_xy=True).transform
    circle = transform(fwd, Point(lon, lat)).buffer(radius_km * KM,
                                                    resolution=64)
    return mapping(transform(back, circle))


def us_county(fips: str) -> dict | None:
    """A US county boundary, from the Census TIGERweb service."""
    if fips in _district_cache:
        return _district_cache[fips]
    try:
        r = S.get(US_COUNTY_SERVICE, params={
            "where": "GEOID = '{}'".format(sql(fips)),
            "outFields": "NAME,GEOID,STATE", "returnGeometry": "true",
            "outSR": 4326, "f": "json"}, timeout=TIMEOUT)
        r.raise_for_status()
        d = r.json()
    except Exception:
        return None
    feats = d.get("features") or []
    if not feats:
        _district_cache[fips] = None
        return None
    rings = (feats[0].get("geometry") or {}).get("rings")
    if not rings:
        _district_cache[fips] = None
        return None
    geom = {"type": "MultiPolygon", "coordinates": [[r] for r in rings]}
    _district_cache[fips] = geom
    return geom


def us_counties_for(text: str) -> tuple[list, str]:
    """The counties implied by a US mill town. Several, where the company
    operates across several - which is the normal case.

    Green Diamond runs across Del Norte, Humboldt and Trinity; Roseburg across
    Douglas and Coos. Returning one county each would have covered about a
    third of the first and missed where the second's wood actually comes from.
    """
    t = re.sub(r"[^A-Z0-9 ]", " ", str(text or "").upper())
    t = re.sub(r"\s+", " ", t).strip()
    for town in sorted(US_TOWN_COUNTY, key=len, reverse=True):
        if town in t or town.replace(" ", "") in t.replace(" ", ""):
            return US_TOWN_COUNTY[town], town
    return [], ""


def national_forest_for(text: str) -> tuple[str, str, str]:
    t = re.sub(r"[^A-Z0-9 ]", " ", str(text or "").upper())
    for town in sorted(NATIONAL_FOREST, key=len, reverse=True):
        if town in t:
            name, region = NATIONAL_FOREST[town]
            return name, region, town
    return "", "", ""


def usfs_forest(name: str) -> dict | None:
    """A national forest boundary, from the Forest Service EDW service."""
    key = "usfs:" + name
    if key in _district_cache:
        return _district_cache[key]
    try:
        r = S.get(USFS_SERVICE, params={
            "where": "FORESTNAME = '{}'".format(sql(name)),
            "outFields": "FORESTNAME,REGION", "returnGeometry": "true",
            "outSR": 4326, "f": "json"}, timeout=TIMEOUT)
        r.raise_for_status()
        d = r.json()
    except Exception:
        return None
    feats = d.get("features") or []
    if not feats:
        _district_cache[key] = None
        return None
    rings = (feats[0].get("geometry") or {}).get("rings")
    if not rings:
        _district_cache[key] = None
        return None
    geom = {"type": "MultiPolygon", "coordinates": [[r] for r in rings]}
    _district_cache[key] = geom
    return geom


# ────────────────────────── method 2: district ─────────────────────────────

_district_cache: dict[str, dict] = {}


def district_polygon(code: str) -> dict | None:
    """A natural resource district boundary. The Domtar pattern: a named
    administrative area, from a published source, rather than a shape we drew.
    """
    if code in _district_cache:
        return _district_cache[code]
    try:
        d = post("{}/{}/query".format(ROOT, DISTRICTS), {
            "where": "ORG_UNIT = '{}'".format(sql(code)),
            "outFields": "DISTRICT_NAME,ORG_UNIT", "returnGeometry": "true",
            "outSR": 4326, "resultRecordCount": 2})
    except Exception:
        return None
    feats = d.get("features") or []
    if not feats:
        _district_cache[code] = None
        return None
    g = feats[0].get("geometry") or {}
    rings = g.get("rings")
    if not rings:
        _district_cache[code] = None
        return None
    geom = {"type": "MultiPolygon", "coordinates": [[r] for r in rings]}
    _district_cache[code] = geom
    return geom


# ───────────────────────── method 1: operator tenure ───────────────────────

def operator_blocks(client_number: str, limit: int = 0,
                    log=None) -> tuple[list[dict], int]:
    """Every cut block this client holds, as features. Returns (blocks, total).

    Paged on OBJECTID. The register caps a page at a thousand and does not
    honour an offset, so an offset loop silently under-reports.

    No cap by default. An earlier version stopped at three thousand blocks per
    supplier, which is under-declaration by accident - the output said "2,459
    blocks" whether that was all of them or a truncated set. If a limit is
    given now, the true total is returned alongside so the shortfall is
    visible rather than invisible.
    """
    log = log or (lambda *_: None)
    out, cursor, guard = [], None, 0
    base = "CLIENT_NUMBER = '{}'".format(sql(client_number))
    try:
        total = int(post("{}/{}/query".format(ROOT, BLOCKS),
                         {"where": base, "returnCountOnly": "true"}
                         ).get("count") or 0)
    except Exception:
        total = 0
    while guard < 400 and (limit <= 0 or len(out) < limit):
        guard += 1
        where = base if cursor is None else \
            "({}) AND OBJECTID > {}".format(base, cursor)
        try:
            d = post("{}/{}/query".format(ROOT, BLOCKS), {
                "where": where,
                "outFields": ("OBJECTID,TIMBER_MARK,CLIENT_NAME,"
                              "CLIENT_NUMBER,"
                              "GEOGRAPHIC_DISTRICT_CODE,FEATURE_AREA,"
                              "DISTURBANCE_START_DATE"),
                "returnGeometry": "true", "outSR": 4326,
                "orderByFields": "OBJECTID", "resultRecordCount": 1000})
        except Exception as exc:
            log("      block fetch failed: {}".format(str(exc)[:70]))
            break
        feats = d.get("features") or []
        if not feats:
            break
        oids = []
        for f in feats:
            a = f.get("attributes", {})
            g = f.get("geometry") or {}
            rings = g.get("rings")
            if not rings:
                continue
            oids.append(a.get("OBJECTID"))
            out.append({"geometry": {"type": "MultiPolygon",
                                     "coordinates": [[r] for r in rings]},
                        "attrs": a})
        oids = [o for o in oids if o is not None]
        if not oids:
            break
        nxt = max(oids)
        if cursor is not None and nxt <= cursor:
            break
        cursor = nxt
        if len(feats) < 1000:
            break
        time.sleep(0.1)
    if limit > 0 and total > len(out):
        log("      {} of {} blocks fetched - the rest were cut off by "
            "--block-limit".format(len(out), total))
    return (out[:limit] if limit > 0 else out), total


# ──────────────────────────────── inputs ───────────────────────────────────

def town_district(text: str) -> tuple[str, str, str]:
    """A district implied by a town name in a source identifier.

    Longest match first, so PORT ALBERNI is not read as a hit on ALBERNI or a
    partial on PORT.
    """
    t = re.sub(r"[^A-Z0-9 ]", " ", str(text or "").upper())
    t = re.sub(r"\s+", " ", t).strip()
    for town in sorted(TOWN_DISTRICT, key=len, reverse=True):
        if town in t or town.replace(" ", "") in t.replace(" ", ""):
            code, name = TOWN_DISTRICT[town]
            return code, name, town
    return "", "", ""


def read_sources(path: str) -> list[dict]:
    """Suppliers and their outstanding sources, from the geometry source map."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = "Suppliers" if "Suppliers" in wb.sheetnames else wb.sheetnames[0]
    rows = list(wb[sheet].iter_rows(values_only=True))
    hdr = [str(h or "").strip().lower() for h in rows[0]]

    def col(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return None

    i_sup, i_code = col("supplier"), col("code")
    i_last = col("weakest system", "last_resort_system")
    i_n = col("sources")
    i_bdt = col("july bdt", "bdt")
    i_jur = col("jurisdiction")
    i_key = col("alias / key in that system", "alias_or_key", "source codes")
    i_sys = col("systems used", "systems")
    i_out = col("outstanding", "outstanding sources")
    out = []
    for r in rows[1:]:
        if i_sup is None or i_sup >= len(r) or not r[i_sup]:
            continue
        out.append({
            "supplier": str(r[i_sup]).strip(),
            "code": str(r[i_code] or "").strip() if i_code is not None else "",
            "weakest": str(r[i_last] or "").strip() if i_last is not None else "",
            "sources": r[i_n] if i_n is not None else 0,
            "bdt": r[i_bdt] if i_bdt is not None else 0,
            "jurisdiction": str(r[i_jur] or "") if i_jur is not None else "",
            "keys": str(r[i_key] or "") if i_key is not None else "",
            "systems": str(r[i_sys] or "") if i_sys is not None else "",
            "outstanding": r[i_out] if i_out is not None else "",
        })
    return out


def read_source_identifiers(path: str) -> dict:
    """Supplier -> the identifiers on their sources.

    The mill town lives here, in the client's own supply record - PARKSVILLE,
    MERRITT, CASTLEGAR. An earlier version read it from the geometry summary
    instead, where the field says "—" for precisely the suppliers that needed
    it, so the town route never fired for any of them.
    """
    out: dict[str, list] = {}
    if not path or not os.path.isfile(path):
        return out
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        return out
    hdr = [str(h or "").strip().upper() for h in rows[0]]

    def col(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return None

    i_sup = col("SUPPID")
    i_uid = col("UNITID")
    i_name = col("NAME")
    if i_sup is None:
        return out
    for r in rows[1:]:
        if i_sup >= len(r) or not r[i_sup]:
            continue
        code = str(r[i_sup]).strip()
        for i in (i_uid, i_name):
            if i is not None and i < len(r) and r[i]:
                out.setdefault(code, []).append(str(r[i]))
    return out


def read_mills(path: str) -> dict:
    out = {}
    if not path or not os.path.isfile(path):
        return out
    with open(path, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            name = (r.get("supplier") or "").strip()
            if name:
                out[name] = r
    return out


# ───────────────────────────────── build ───────────────────────────────────

# Washington suppliers resolve through the state's Forest Practices
# Application register, by company name. That is a different pipeline, not a
# missing catchment, and counting them as gaps put 14,388 BDT in the wrong
# column.
US_REGISTER = {
    "WEYER", "SIERRA", "MANKE", "INTERFOR", "HERMANN", "HERMAN",
    "WILLIS", "ALTA", "GREEN",
}


# US mill towns. The client's own jurisdiction field is not reliable for this
# - Alta Forest Products is flagged BC in SOURCE.xlsx while milling at Port
# Angeles and Tacoma, which are both Washington. Where the record and the town
# disagree, the town is the better evidence.
# US mill towns, and the county each sits in. A county is the administrative
# area a US state publishes harvest activity against, and it is the same
# pattern Domtar uses: a named place joined to a published boundary.
#
# The county is where the mill is, not where the wood grew. It bounds a
# search. Where a supplier also appears in a state's harvest register - as all
# the Washington ones do - the register is the better answer and this is the
# fallback.
US_TOWN_COUNTY = {
    # Washington. Every one of these suppliers also appears in the state's
    # Forest Practices Application register, which is a far better answer -
    # these counties are the floor beneath that.
    "PORT ANGELES": [("53009", "Clallam", "WA")],
    "ABERDEEN": [("53027", "Grays Harbor", "WA")],
    "HOQUIAM": [("53027", "Grays Harbor", "WA")],
    "TACOMA": [("53053", "Pierce", "WA")],
    "SHELTON": [("53045", "Mason", "WA")],
    "BURLINGTON": [("53057", "Skagit", "WA")],
    "MOUNT VERNON": [("53057", "Skagit", "WA")],
    "MORTON": [("53041", "Lewis", "WA")],
    "CENTRALIA": [("53041", "Lewis", "WA")],
    "CHEHALIS": [("53041", "Lewis", "WA")],
    "LONGVIEW": [("53015", "Cowlitz", "WA")],
    "OLYMPIA": [("53067", "Thurston", "WA")],
    "EVERETT": [("53061", "Snohomish", "WA")],
    "DARRINGTON": [("53061", "Snohomish", "WA")],
    "BELLINGHAM": [("53073", "Whatcom", "WA")],
    "RAYMOND": [("53049", "Pacific", "WA")],
    "FORKS": [("53009", "Clallam", "WA")],

    # Viking Lumber mills at Klawock, near Craig. Confirmed August 2026: their
    # logs come from Tongass National Forest timber sales, Alaska Mental
    # Health Trust land and state sales. The census area is the administrative
    # answer; the Tongass boundary below is the better one.
    "CRAIG": [("02198", "Prince of Wales-Hyder", "AK")],
    "KLAWOCK": [("02198", "Prince of Wales-Hyder", "AK")],

    # Green Diamond's chip facility is at Samoa on Humboldt Bay, but their
    # timberland runs to roughly 428,000 acres across three counties.
    # Confirmed August 2026 against their own habitat conservation plans.
    # Humboldt alone would cover about a third of it.
    "EUREKA": [("06023", "Humboldt", "CA"),
               ("06015", "Del Norte", "CA"),
               ("06105", "Trinity", "CA")],
    "SAMOA": [("06023", "Humboldt", "CA"),
              ("06015", "Del Norte", "CA"),
              ("06105", "Trinity", "CA")],
    "KORBEL": [("06023", "Humboldt", "CA"),
               ("06015", "Del Norte", "CA"),
               ("06105", "Trinity", "CA")],

    # Coos Bay is Roseburg's chip EXPORT TERMINAL, not a mill. Confirmed
    # August 2026. The chips are made at Coquille, Dillard and Riddle, and the
    # company's Oregon timberland sits in Douglas and Coos counties - so the
    # terminal's county alone would miss where the wood actually comes from.
    "COOS BAY": [("41011", "Coos", "OR"), ("41019", "Douglas", "OR")],
    "COUSEBAY": [("41011", "Coos", "OR"), ("41019", "Douglas", "OR")],
    "COQUILLE": [("41011", "Coos", "OR"), ("41019", "Douglas", "OR")],
    "ROSEBURG": [("41019", "Douglas", "OR"), ("41011", "Coos", "OR")],
    "DILLARD": [("41019", "Douglas", "OR")],
    "RIDDLE": [("41019", "Douglas", "OR")],
}

# Where a national forest is the actual source, its own boundary beats a
# county. Viking buys most of its timber from Tongass sales, and the Tongass
# is published as a mapped administrative forest.
NATIONAL_FOREST = {
    "CRAIG": ("Tongass National Forest", "Alaska Region"),
    "KLAWOCK": ("Tongass National Forest", "Alaska Region"),
}

USFS_SERVICE = ("https://apps.fs.usda.gov/arcx/rest/services/EDW"
                "/EDW_ForestSystemBoundaries_01/MapServer/1/query")

# US Census county boundaries, published and stable.
US_COUNTY_SERVICE = ("https://tigerweb.geo.census.gov/arcgis/rest/services"
                     "/TIGERweb/State_County/MapServer/13/query")

US_TOWNS = {
    "PORT ANGELES", "TACOMA", "ABERDEEN", "HOQUIAM", "SHELTON", "MORTON",
    "BURLINGTON", "LONGVIEW", "OLYMPIA", "CENTRALIA", "MOUNT VERNON",
    "EVERETT", "BELLINGHAM", "CHEHALIS", "RAYMOND", "FORKS", "DARRINGTON",
}


def has_us_route(code: str, jurisdiction: str, identifiers: str = "") -> bool:
    """Does this supplier resolve through a US state register?

    Decided on the company name and the mill town together. Trusting the
    jurisdiction field alone left Alta Forest Products - 777 BDT, milling at
    Port Angeles - counted as a British Columbia supplier with no catchment.
    """
    named = any(k in (code or "").upper() for k in US_REGISTER)
    if not named:
        return False
    text = str(identifiers or "").upper()
    if any(t in text for t in US_TOWNS):
        return True
    jur = jurisdiction or ""
    return not ("British Columbia" in jur and "Washington" not in jur)


def feature(geom, props) -> dict:
    return {"type": "Feature", "geometry": geom, "properties": props}


def build(sources, mills, aliases, radius_km, block_limit, identifiers=None,
          stated=None, log=print) -> dict:
    feats, summary = [], []
    identifiers = identifiers or {}
    stated = stated or {}

    for s in sources:
        name = s["supplier"]
        weakest = s["weakest"]
        # Sources already carrying real harvest geometry need no catchment.
        # Drawing one anyway would declare the same fibre twice, once
        # precisely and once as an assumption.
        # Only a genuinely-held system means no catchment is needed. An
        # earlier version also treated "FTEN client tenure" as held, because
        # WWK and WWW carry it on resolved sources - but the same string is
        # used for outstanding sources routed there, and reading it as held
        # skipped Aspen Planers and twenty others, removing 24,806 BDT of
        # catchment. WWK and WWW appear as gaps instead, which is cosmetic and
        # nil volume. Over-building a catchment is recoverable; missing one is
        # not.
        if weakest in ("FTEN cut block", "ParcelMap BC", "Harmac LIMS"):
            continue
        if "internal" in name.lower() or "mill yard" in name.lower():
            continue
        ids_text = " ".join(identifiers.get(s["code"], []))
        us_named = has_us_route(s["code"], s.get("jurisdiction", ""), ids_text)
        counties, us_town = us_counties_for(ids_text)
        forest, region, f_town = national_forest_for(ids_text)

        if us_named or counties or forest:
            made_us, parts = 0, []
            # A national forest, where that is where the wood actually comes
            # from. Viking buys most of its timber from Tongass sales, so the
            # forest boundary is a truer catchment than the census area.
            if forest:
                geom = usfs_forest(forest)
                if geom:
                    feats.append(feature(geom, {
                        "harp_supplier": name,
                        "harp_supplier_code": s["code"],
                        "harp_method": "national forest",
                        "harp_source_system": "USDA Forest Service boundaries",
                        "harp_key": forest, "harp_key_name": forest,
                        "harp_basis": ("this supplier buys its timber from {} "
                                       "sales; mill at '{}'".format(forest,
                                                                    f_town)),
                        "harp_declared_by_supplier": False,
                        "harp_tier": "P3a", "harp_is_envelope": True,
                        "harp_traceability": "inferred",
                        "harp_note": ("the national forest the supplier buys "
                                      "from. Federal land, and a truer "
                                      "catchment than the census area"),
                    }))
                    made_us += 1
                    parts.append(forest)

            for fips, county, state in counties:
                geom = us_county(fips)
                if not geom:
                    continue
                feats.append(feature(geom, {
                    "harp_supplier": name, "harp_supplier_code": s["code"],
                    "harp_method": "named county",
                    "harp_source_system": "US Census county boundaries",
                    "harp_key": fips,
                    "harp_key_name": "{} County, {}".format(county, state),
                    "harp_basis": ("inferred from the mill town '{}' in the "
                                   "source identifier".format(us_town)),
                    "harp_declared_by_supplier": False,
                    "harp_also_in_register": bool(us_named),
                    "harp_tier": "P3a", "harp_is_envelope": True,
                    "harp_traceability": "inferred",
                    "harp_note": ("one of {} counties this supplier operates "
                                  "across".format(len(counties))
                                  if len(counties) > 1 else
                                  "the county the mill sits in"),
                }))
                made_us += 1
                parts.append("{} Co, {}".format(county, state))

            m = ("national forest" if forest and made_us else
                 "named county" if made_us else "US register")
            n = (", ".join(parts) + ("  (also in FPARS)" if us_named else "")
                 if made_us else "in FPARS, no mill town to place")
            summary.append({"supplier": name, "code": s["code"],
                            "jurisdiction": s["jurisdiction"],
                            "sources": s["sources"], "bdt": s["bdt"],
                            "method": m, "features": made_us, "note": n})
            log("  {:<32} {:<18} {}".format(name[:32], m, n[:44]))
            continue

        # Set up for the BC routes below. These were lost when the US branch
        # was spliced in above, which is the sort of thing a splice does.
        mill = mills.get(name, {})
        accepted = aliases.for_supplier(name) if aliases else []
        method, note, made = "", "", 0

        # 1 — the operator's own tenure
        if accepted:
            for a in accepted:
                blocks, held = operator_blocks(a.client_number, block_limit, log)
                for b in blocks:
                    at = b["attrs"]
                    feats.append(feature(b["geometry"], {
                        "harp_supplier": name, "harp_supplier_code": s["code"],
                        "harp_method": "operator tenure",
                        # The block's own holder, not the alias we matched on.
                        # They agree in the normal case, and where they differ
                        # the register is right.
                        "ProducerName": (at.get("CLIENT_NAME")
                                         or a.client_name or ""),
                        "harp_producer_number": (at.get("CLIENT_NUMBER")
                                                 or a.client_number or ""),
                        "harp_producer_source": "forest register",
                        "harp_source_system": "FTEN cut block register",
                        "harp_key": a.client_number,
                        "harp_key_name": a.client_name or at.get("CLIENT_NAME"),
                        "harp_alias_decided_by": a.decided_by,
                        "harp_alias_reason": a.reason,
                        "timber_mark": at.get("TIMBER_MARK"),
                        "district": at.get("GEOGRAPHIC_DISTRICT_CODE"),
                        "area_ha": round(float(at.get("FEATURE_AREA") or 0)
                                         / 10000.0, 2),
                        "harp_tier": "P2a", "harp_is_envelope": True,
                        "harp_traceability": "inferred",
                        "harp_note": ("everywhere this operator cut, not what "
                                      "this client bought"),
                        "harp_attribution": ATTRIBUTION,
                    }))
                    made += 1
            if made:
                method = "operator tenure"
                note = "{} blocks from {} accepted alias(es)".format(
                    made, len(accepted))

        # 2 — the district the mill sits in
        if not made and str(mill.get("district_code") or "").strip():
            code = mill["district_code"].strip()
            geom = district_polygon(code)
            if geom:
                feats.append(feature(geom, {
                    "harp_supplier": name, "harp_supplier_code": s["code"],
                    "harp_method": "named district",
                    "harp_source_system": "BC Natural Resource Districts",
                    "harp_key": code,
                    "harp_key_name": mill.get("district", ""),
                    "harp_basis": mill.get("how_established", ""),
                    "harp_declared_by_supplier": False,
                    "harp_mill": mill.get("facility", ""),
                    "harp_tier": "P3a", "harp_is_envelope": True,
                    "harp_traceability": "inferred",
                    "harp_note": ("the district the mill sits in. The wood may "
                                  "come from elsewhere - this bounds a search, "
                                  "it does not locate a harvest"),
                    "harp_attribution": ATTRIBUTION,
                }))
                made, method = 1, "named district"
                note = "district {}".format(code)

        # 2b — a district implied by the mill town in the source identifier
        if not made:
            text = " ".join(identifiers.get(s["code"], [])) or s.get("keys", "")
            code, dname, town = town_district(text)
            if code:
                geom = district_polygon(code)
                if geom:
                    feats.append(feature(geom, {
                        "harp_supplier": name,
                        "harp_supplier_code": s["code"],
                        "harp_method": "named district",
                        "harp_source_system": "BC Natural Resource Districts",
                        "harp_key": code, "harp_key_name": dname,
                        "harp_basis": ("inferred from the mill town '{}' in "
                                       "the source identifier".format(town)),
                        "harp_declared_by_supplier": False,
                        "harp_tier": "P3a", "harp_is_envelope": True,
                        "harp_traceability": "inferred",
                        "harp_note": ("the district the mill sits in, inferred "
                                      "rather than declared. The wood may come "
                                      "from elsewhere - Mercer mills at "
                                      "Castlegar and draws from across the "
                                      "Kootenays"),
                        "harp_attribution": ATTRIBUTION,
                    }))
                    made, method = 1, "named district"
                    note = "district {} from '{}'".format(code, town)

        # 2b — an area somebody stated by hand
        #
        # After every register route and before the mill buffer. A stated
        # area never overrides a tenure record, and always beats a circle
        # drawn round a mill - which is a guess with no author, where this at
        # least has one.
        if not made and name in stated:
            e = stated[name]
            parts = []
            for code in e.get("districts") or []:
                geom = district_polygon(code)
                if not geom:
                    continue
                feats.append(feature(geom, {
                    "harp_supplier": name,
                    "harp_supplier_code": s["code"],
                    "harp_method": "stated area",
                    "harp_source_system": "BC Natural Resource Districts",
                    "harp_key": code,
                    "harp_basis": "{} on {}: {}".format(
                        e.get("stated_by") or "?", e.get("stated_at") or "?",
                        e.get("basis") or "stated"),
                    # Only a supplier's own words count as declared. Anyone
                    # else stating it, however well informed, is inference
                    # with an author.
                    "harp_declared_by_supplier": e.get("basis") == "supplier",
                    "harp_tier": "P3a",
                    "harp_is_envelope": True,
                    "harp_traceability": "inferred",
                    "harp_note": (e.get("note") or
                                  "an operating area stated by hand, because "
                                  "no register could place this supplier"),
                    "harp_attribution": ATTRIBUTION,
                }))
                parts.append(code)
            for fips in e.get("counties") or []:
                geom = us_county(fips)
                if not geom:
                    continue
                feats.append(feature(geom, {
                    "harp_supplier": name,
                    "harp_supplier_code": s["code"],
                    "harp_method": "stated area",
                    "harp_source_system": "US Census county boundaries",
                    "harp_key": fips,
                    "harp_basis": "{} on {}: {}".format(
                        e.get("stated_by") or "?", e.get("stated_at") or "?",
                        e.get("basis") or "stated"),
                    "harp_declared_by_supplier": e.get("basis") == "supplier",
                    "harp_tier": "P3a",
                    "harp_is_envelope": True,
                    "harp_traceability": "inferred",
                    "harp_note": (e.get("note") or
                                  "an operating area stated by hand"),
                    "harp_attribution": ATTRIBUTION,
                }))
                parts.append(fips)
            if parts:
                made, method = len(parts), "stated area"
                note = "{} stated by {}".format(", ".join(parts),
                                                e.get("stated_by") or "?")

        # 3 — a circle round the mill
        if not made:
            try:
                lat = float(mill.get("latitude") or "")
                lon = float(mill.get("longitude") or "")
            except (TypeError, ValueError):
                lat = lon = None
            if lat and lon:
                geom = mill_buffer(lat, lon, radius_km)
                if geom:
                    feats.append(feature(geom, {
                        "harp_supplier": name,
                        "harp_supplier_code": s["code"],
                        "harp_method": "mill buffer",
                        "harp_source_system": "BC Major Timber Processing "
                                              "Facilities",
                        "harp_key": mill.get("facility", ""),
                        "harp_radius_km": radius_km,
                        "harp_mill_lat": lat, "harp_mill_lon": lon,
                        "harp_tier": "P3a", "harp_is_envelope": True,
                        "harp_traceability": "inferred",
                        "harp_note": ("a circle of assumed haul distance. The "
                                      "radius is a starting position, not a "
                                      "measurement"),
                        "harp_attribution": ATTRIBUTION,
                    }))
                    made, method = 1, "mill buffer"
                    note = "{:.0f} km around {}".format(
                        radius_km, mill.get("city") or "the mill")

        # 4 — nothing
        if not made:
            method = "none"
            note = "no tenure, no district, no mill point"

        summary.append({"supplier": name, "code": s["code"],
                        "jurisdiction": s["jurisdiction"],
                        "sources": s["sources"], "bdt": s["bdt"],
                        "method": method, "features": made, "note": note})
        log("  {:<32} {:<18} {}".format(name[:32], method, note[:44]))

    return {"type": "FeatureCollection", "name": "harp_catchments",
            "metadata": {
                "generated": datetime.now().isoformat(timespec="seconds"),
                "radius_km": radius_km,
                "attribution": ATTRIBUTION,
                "note": ("Search areas, not harvest areas. Every polygon here "
                         "bounds a region within which a supplier's fibre "
                         "plausibly originated. Change detection finds the "
                         "ground actually disturbed; nothing here is a plot "
                         "traced."),
                "methods": {
                    "operator tenure": "the company's own cut blocks, FTEN",
                    "named district": "published district boundary",
                    "mill buffer": "circle of assumed haul distance",
                }},
            "features": feats}, summary
